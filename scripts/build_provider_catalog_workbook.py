"""
Build a docs-only provider catalog workbook that compares Alterra and Kiosbank.

Usage:
  py -3 scripts/build_provider_catalog_workbook.py

Reads:
  docs/ppob/alterra/20262801 - BPA Product Catalogue  (OPEN).xlsx
  docs/ppob/kiosbank/Pricelist Kiosbank.xlsx

Writes:
  docs/ppob/[GTD] Provider Catalog Workbook.xlsx
"""

from __future__ import annotations

import argparse
import os
import re
import sys
from pathlib import Path
from collections import defaultdict
from datetime import datetime

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


SCRIPT_DIR = Path(__file__).resolve().parent
if str(SCRIPT_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPT_DIR))

import parse_alterra_catalog as alterra  # noqa: E402
import parse_kiosbank_pricelist as kiosbank  # noqa: E402


ROOT_DIR = SCRIPT_DIR.parent
ALTERRA_XLSX = ROOT_DIR / "docs" / "ppob" / "alterra" / "20262801 - BPA Product Catalogue  (OPEN).xlsx"
KIOSBANK_XLSX = ROOT_DIR / "docs" / "ppob" / "kiosbank" / "Pricelist Kiosbank.xlsx"
DEFAULT_OUTPUT = ROOT_DIR / "docs" / "ppob" / "[GTD] Provider Catalog Workbook.xlsx"

PREPAID_CATEGORY_ORDER = [
    "Pulsa",
    "Paket Data",
    "Listrik",
    "E-Money",
    "Voucher",
    "Voucher Game",
    "Streaming",
    "TV Kabel",
    "Gas PGN",
    "Donasi",
    "Tiket",
    "Properti",
    "Edukasi",
]

POSTPAID_CATEGORY_ORDER = [
    "Listrik",
    "BPJS Kesehatan",
    "Telepon Pascabayar",
    "Internet",
    "TV Kabel",
    "Gas PGN",
    "PDAM",
    "PBB",
    "Multifinance",
    "Properti",
    "Edukasi",
]

PREPAID_COLUMNS = [
    "kode_produk_kita",
    "jenis_produk",
    "brand",
    "nama_sku_kita",
    "match_status",
    "kiosbank_sheet",
    "kiosbank_kode_produk",
    "kiosbank_nama_produk_raw",
    "kiosbank_harga_jual",
    "kiosbank_fee_mitra_raw",
    "kiosbank_fee_mitra_numeric",
    "alterra_sheet",
    "alterra_product_id",
    "alterra_nama_produk_raw",
    "alterra_price",
    "alterra_commission",
    "catatan",
]

POSTPAID_COLUMNS = [
    "kode_produk_kita",
    "jenis_produk",
    "brand",
    "nama_sku_kita",
    "match_status",
    "kiosbank_sheet",
    "kiosbank_kode_produk",
    "kiosbank_nama_biller_raw",
    "kiosbank_biaya_admin",
    "kiosbank_fee_mitra",
    "alterra_sheet",
    "alterra_product_id",
    "alterra_nama_produk_raw",
    "alterra_service_fee",
    "alterra_collection_fee_for_partner",
    "catatan",
]

NO_MATCH_CATEGORIES = {"PDAM", "PBB", "Multifinance"}


def note_join(*notes: str) -> str:
    seen = set()
    ordered = []
    for note in notes:
        if not note:
            continue
        for part in str(note).split(";"):
            item = part.strip()
            if not item or item in seen:
                continue
            seen.add(item)
            ordered.append(item)
    return "; ".join(ordered)


def safe_string(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalize_match_name(value: str) -> str:
    text = safe_string(value).lower()
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def names_clearly_related(left: str, right: str) -> bool:
    a = normalize_match_name(left)
    b = normalize_match_name(right)
    if not a or not b:
        return False
    if a == b:
        return True
    if len(a) >= 6 and a in b:
        return True
    if len(b) >= 6 and b in a:
        return True

    tokens_a = set(a.split())
    tokens_b = set(b.split())
    if len(tokens_a) >= 2 and tokens_a.issubset(tokens_b):
        return True
    if len(tokens_b) >= 2 and tokens_b.issubset(tokens_a):
        return True
    return False


def extract_conservative_denom(value: str) -> int:
    denom = kiosbank.extract_denom(value)
    if denom > 0:
        return denom

    numbers = re.findall(r"\d[\d.]*", safe_string(value))
    parsed = []
    for item in numbers:
        cleaned = item.replace(".", "")
        if not cleaned:
            continue
        try:
            parsed.append(int(cleaned))
        except ValueError:
            continue

    unique = []
    seen = set()
    for item in parsed:
        if item <= 0 or item in seen:
            continue
        seen.add(item)
        unique.append(item)

    if len(unique) == 1:
        return unique[0]
    return 0


def source_sheet_for_type(product_type: str) -> str:
    return "Prepaid" if product_type == "prepaid" else "Postpaid"


def category_sort_rank(sheet_name: str, category: str) -> tuple[int, str]:
    order = PREPAID_CATEGORY_ORDER if sheet_name == "Prepaid" else POSTPAID_CATEGORY_ORDER
    try:
        return (order.index(category), "")
    except ValueError:
        return (len(order), safe_string(category).lower())


def numeric_sort_value(*values) -> int:
    for value in values:
        if value is None:
            continue
        if isinstance(value, (int, float)) and int(value) > 0:
            return int(value)
    return 10**12


def row_sort_amount(sheet_name: str, row: dict) -> int:
    denom = extract_conservative_denom(
        row.get("nama_sku_kita")
        or row.get("kiosbank_nama_produk_raw")
        or row.get("kiosbank_nama_biller_raw")
        or row.get("alterra_nama_produk_raw")
    )
    if denom > 0:
        return denom

    if sheet_name == "Prepaid":
        return numeric_sort_value(row.get("kiosbank_harga_jual"), row.get("alterra_price"))

    return numeric_sort_value(
        row.get("kiosbank_biaya_admin"),
        row.get("alterra_service_fee"),
        row.get("kiosbank_fee_mitra"),
        row.get("alterra_collection_fee_for_partner"),
    )


def build_alterra_records() -> list[dict]:
    workbook = openpyxl.load_workbook(ALTERRA_XLSX, read_only=True)
    sheet_parsers = [
        ("Pulsa", alterra.parse_pulsa),
        ("PLN, BPJS & Postpaid", alterra.parse_pln_bpjs_postpaid),
        ("PBB", alterra.parse_pbb),
        ("Gas & PDAM", alterra.parse_gas_pdam),
        ("Ticket & Ewallet", alterra.parse_ticket_ewallet),
        ("Game", alterra.parse_game),
        ("Streaming & TV", alterra.parse_streaming_tv),
        ("Voucher Deals", alterra.parse_voucher_deals),
        ("Donation", alterra.parse_donation),
        ("Property", alterra.parse_property),
        ("Edukasi", alterra.parse_edukasi),
        ("Voucher Edukasi", alterra.parse_voucher_edukasi),
    ]

    seen_ids = set()
    records: list[dict] = []

    for sheet_name, parser in sheet_parsers:
        if sheet_name not in workbook.sheetnames:
            continue
        for item in parser(workbook[sheet_name]):
            product_id = item["alterra_product_id"]
            if product_id in seen_ids:
                continue
            seen_ids.add(product_id)
            raw_name = item.get("description") or item.get("name") or ""
            records.append(
                {
                    "sheet": sheet_name,
                    "type": item["type"],
                    "category": item["category"],
                    "brand": item["brand"],
                    "internal_sku_code": item["sku_code"],
                    "internal_name": item["name"],
                    "raw_name": raw_name,
                    "alterra_product_id": int(product_id),
                    "price": int(item.get("price") or 0),
                    "admin": int(item.get("admin") or 0),
                    "commission": int(item.get("commission") or 0),
                }
            )

    workbook.close()
    return records


def build_alterra_lookup(records: list[dict]) -> dict:
    index = {
        "by_name": defaultdict(set),
        "by_denom": defaultdict(set),
        "by_group": defaultdict(set),
        "by_sku": {},
    }
    for record in records:
        category = record["category"].lower()
        brand = record["brand"].lower()
        product_type = record["type"]
        names = [record["internal_name"], record["raw_name"]]

        index["by_sku"][record["internal_sku_code"]] = record
        index["by_group"][(category, brand, product_type)].add(record["internal_sku_code"])

        for raw_name in names:
            name = normalize_match_name(raw_name)
            if not name:
                continue
            index["by_name"][(category, brand, product_type, name)].add(record["internal_sku_code"])
            denom = extract_conservative_denom(raw_name)
            if denom > 0:
                index["by_denom"][(category, brand, product_type, denom)].add(record["internal_sku_code"])
    return index


def only_match(candidates: set[str]) -> str | None:
    if len(candidates) != 1:
        return None
    return next(iter(candidates))


def find_conservative_match(alterra_lookup: dict, record: dict) -> str | None:
    category = record["category"].lower()
    brand = record["brand"].lower()
    product_type = record["type"]
    raw_name = record["raw_name"]

    normalized_name = normalize_match_name(raw_name)
    if normalized_name:
        match = only_match(
            alterra_lookup["by_name"].get((category, brand, product_type, normalized_name), set())
        )
        if match:
            return match

    denom = extract_conservative_denom(raw_name)
    if denom > 0:
        match = only_match(
            alterra_lookup["by_denom"].get((category, brand, product_type, denom), set())
        )
        if match:
            return match

    group_match = only_match(alterra_lookup["by_group"].get((category, brand, product_type), set()))
    if not group_match:
        return None

    candidate = alterra_lookup["by_sku"][group_match]
    if names_clearly_related(record["raw_name"], candidate["raw_name"]) or names_clearly_related(
        record["raw_name"], candidate["internal_name"]
    ):
        return group_match
    return None


def parse_kiosbank_postpaid_sheet(ws, category, brand_from_name=False, fixed_brand=None):
    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not kiosbank.is_data_row(row):
            continue

        code = safe_string(row[1])
        raw_name = safe_string(row[2]) or code
        admin = kiosbank.safe_int(row[3])
        fee_raw = row[4]
        commission = kiosbank.parse_fee(fee_raw)

        if fixed_brand:
            brand = fixed_brand
        elif brand_from_name:
            brand = kiosbank.normalize_brand(raw_name)
        else:
            brand = kiosbank.normalize_brand(category)

        records.append(
            {
                "sheet": ws.title,
                "kb_code": code,
                "raw_name": raw_name,
                "category": category,
                "brand": brand,
                "type": "postpaid",
                "price": 0,
                "admin": admin,
                "commission": commission,
                "fee_raw": fee_raw,
                "note": "",
            }
        )
    return records


def parse_kiosbank_pln(ws):
    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not kiosbank.is_data_row(row):
            continue

        code = safe_string(row[1])
        raw_name = safe_string(row[2]) or code
        admin = kiosbank.safe_int(row[3])
        fee_raw = row[4]
        commission = kiosbank.parse_fee(fee_raw)
        name_lower = raw_name.lower()
        is_prepaid = "prepaid" in name_lower
        note = ""
        if is_prepaid:
            note = note_join(note, f"kiosbank prepaid source uses biaya_admin={admin}")

        records.append(
            {
                "sheet": ws.title,
                "kb_code": code,
                "raw_name": raw_name,
                "category": "Listrik",
                "brand": "PLN",
                "type": "prepaid" if is_prepaid else "postpaid",
                "price": 0,
                "admin": admin,
                "commission": commission,
                "fee_raw": fee_raw,
                "note": note,
            }
        )
    return records


def parse_kiosbank_pgn(ws):
    records = []
    section = "postpaid"
    for row in ws.iter_rows(min_row=2, values_only=True):
        label = safe_string(row[0]).upper()
        if "PREPAID" in label:
            section = "prepaid"
            continue
        if "POSTPAID" in label:
            section = "postpaid"
            continue
        if not kiosbank.is_data_row(row):
            continue

        code = safe_string(row[1])
        raw_name = safe_string(row[2]) or code
        admin = kiosbank.safe_int(row[3])
        fee_raw = row[4]
        commission = kiosbank.parse_fee(fee_raw)
        note = ""
        if section == "prepaid":
            note = note_join(note, f"kiosbank prepaid source uses biaya_admin={admin}")

        records.append(
            {
                "sheet": ws.title,
                "kb_code": code,
                "raw_name": raw_name,
                "category": "Gas PGN",
                "brand": "PGN",
                "type": section,
                "price": 0,
                "admin": admin,
                "commission": commission,
                "fee_raw": fee_raw,
                "note": note,
            }
        )
    return records


def parse_kiosbank_prepaid_sheet(
    ws,
    category,
    brand_from_sections=False,
    fixed_brand=None,
    price_col=3,
    fee_col=4,
    name_col=2,
    code_col=1,
):
    records = []
    current_brand = fixed_brand or "LAINNYA"

    for row in ws.iter_rows(min_row=1, values_only=True):
        if brand_from_sections and row[code_col] is None and row[0] is not None:
            header = safe_string(row[0])
            if header and not header[0].isdigit() and not header.lower().startswith(("no", "ketentuan", "1.", "2.")):
                current_brand = kiosbank.normalize_brand(header)
                continue

        if row[code_col] is None:
            continue

        code = safe_string(row[code_col])
        if code.lower() in ("kode produk", "id product", "no.", "no"):
            continue

        raw_name = safe_string(row[name_col]) or code
        price = kiosbank.safe_int(row[price_col])
        fee_raw = row[fee_col]
        commission = kiosbank.parse_fee(fee_raw)
        if price <= 0:
            continue

        if not brand_from_sections and not fixed_brand:
            brand = kiosbank.normalize_brand(raw_name.split(" ")[0] if " " in raw_name else raw_name)
        else:
            brand = current_brand

        note = ""
        if isinstance(fee_raw, str) and fee_raw.strip().lower().startswith("markup"):
            note = note_join(note, "Markup By Mitra")

        records.append(
            {
                "sheet": ws.title,
                "kb_code": code,
                "raw_name": raw_name,
                "category": category,
                "brand": brand,
                "type": "prepaid",
                "price": price,
                "admin": 0,
                "commission": commission,
                "fee_raw": fee_raw,
                "note": note,
            }
        )
    return records


def parse_kiosbank_uang_elektronik(ws):
    records = []
    for row in ws.iter_rows(min_row=1, values_only=True):
        if row[0] is None:
            continue
        code = safe_string(row[0])
        if code.lower() in ("id product", "biller", ""):
            continue

        try:
            int(code)
        except ValueError:
            continue

        raw_name = safe_string(row[1]) or code
        price = kiosbank.safe_int(row[2])
        fee_raw = row[3]
        commission = kiosbank.parse_fee(fee_raw)
        if price <= 0:
            continue

        note = ""
        if isinstance(fee_raw, str) and fee_raw.strip().lower().startswith("markup"):
            note = note_join(note, "Markup By Mitra")

        records.append(
            {
                "sheet": ws.title,
                "kb_code": code,
                "raw_name": raw_name,
                "category": "E-Money",
                "brand": kiosbank.normalize_brand(raw_name),
                "type": "prepaid",
                "price": price,
                "admin": 0,
                "commission": commission,
                "fee_raw": fee_raw,
                "note": note,
            }
        )
    return records


def parse_kiosbank_nexparabola(ws):
    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not kiosbank.is_data_row(row):
            continue

        code = safe_string(row[1])
        raw_name = safe_string(row[2]) or code
        price = kiosbank.safe_int(row[3])
        fee_raw = row[4]
        commission = kiosbank.parse_fee(fee_raw)
        if price <= 0:
            continue

        records.append(
            {
                "sheet": ws.title,
                "kb_code": code,
                "raw_name": raw_name,
                "category": "TV Kabel",
                "brand": "NEX PARABOLA",
                "type": "prepaid",
                "price": price,
                "admin": 0,
                "commission": commission,
                "fee_raw": fee_raw,
                "note": "",
            }
        )
    return records


def build_kiosbank_records() -> list[dict]:
    workbook = openpyxl.load_workbook(KIOSBANK_XLSX, read_only=True)
    records = []

    postpaid_sheets = [
        ("PDAM", "PDAM", False, "PDAM"),
        ("PBB", "PBB", False, "PBB"),
        ("MULTIFINANCE", "Multifinance", True, None),
        ("BPJS", "BPJS Kesehatan", False, "BPJS"),
        ("INTERNET&TV KABEL", "Internet", True, None),
        ("TELEKOMUNIKASI", "Telepon Pascabayar", True, None),
        ("TV BERBAYAR", "TV Kabel", True, None),
    ]

    for sheet_name, category, brand_from_name, fixed_brand in postpaid_sheets:
        if sheet_name in workbook.sheetnames:
            records.extend(
                parse_kiosbank_postpaid_sheet(
                    workbook[sheet_name], category, brand_from_name, fixed_brand
                )
            )

    if "PLN" in workbook.sheetnames:
        records.extend(parse_kiosbank_pln(workbook["PLN"]))
    if "PGN" in workbook.sheetnames:
        records.extend(parse_kiosbank_pgn(workbook["PGN"]))
    if "PULSA" in workbook.sheetnames:
        records.extend(parse_kiosbank_prepaid_sheet(workbook["PULSA"], "Pulsa", brand_from_sections=True))
    if "PAKET DATA" in workbook.sheetnames:
        records.extend(
            parse_kiosbank_prepaid_sheet(workbook["PAKET DATA"], "Paket Data", brand_from_sections=True)
        )
    if "VOUCHER GAME" in workbook.sheetnames:
        records.extend(
            parse_kiosbank_prepaid_sheet(workbook["VOUCHER GAME"], "Voucher Game", brand_from_sections=True)
        )
    if "VOUCHER BELANJA" in workbook.sheetnames:
        records.extend(parse_kiosbank_prepaid_sheet(workbook["VOUCHER BELANJA"], "Voucher"))
    if "UANG ELEKTRONIK" in workbook.sheetnames:
        records.extend(parse_kiosbank_uang_elektronik(workbook["UANG ELEKTRONIK"]))
    if "NEXPARABOLA" in workbook.sheetnames:
        records.extend(parse_kiosbank_nexparabola(workbook["NEXPARABOLA"]))

    workbook.close()
    return records


def assign_kiosbank_internal_codes(kiosbank_records: list[dict], alterra_lookup: dict):
    seen_generated = set()
    for record in kiosbank_records:
        category = record["category"]
        brand = record["brand"]
        if category in NO_MATCH_CATEGORIES:
            matched_sku = None
        else:
            matched_sku = find_conservative_match(alterra_lookup, record)

        if matched_sku:
            record["internal_sku_code"] = matched_sku
            record["matched_to_alterra"] = True
        else:
            cat_slug = re.sub(r"[^A-Z0-9]", "", category.upper())[:10]
            brand_slug = re.sub(r"[^A-Z0-9]", "", brand.upper())[:10]
            generated = f"KB-{cat_slug}-{brand_slug}-{record['kb_code']}"
            while generated in seen_generated:
                generated = f"{generated}-{record['kb_code']}"
            seen_generated.add(generated)
            record["internal_sku_code"] = generated
            record["matched_to_alterra"] = False
            record["note"] = note_join(record.get("note", ""), "unmatched conservative rule")


def build_rows(alterra_records: list[dict], kiosbank_records: list[dict]) -> dict[str, list[dict]]:
    rows = {"Prepaid": [], "Postpaid": []}
    by_sku: dict[str, dict] = {}

    def ensure_row(sheet_name: str, internal_sku_code: str, category: str, brand: str, internal_name: str) -> dict:
        row = by_sku.get(internal_sku_code)
        if row is not None:
            return row

        row = {
            "sheet_name": sheet_name,
            "kode_produk_kita": internal_sku_code,
            "jenis_produk": category,
            "brand": brand,
            "nama_sku_kita": internal_name,
            "match_status": "",
            "kiosbank_sheet": "",
            "kiosbank_kode_produk": "",
            "kiosbank_nama_produk_raw": "",
            "kiosbank_harga_jual": None,
            "kiosbank_fee_mitra_raw": "",
            "kiosbank_fee_mitra_numeric": None,
            "kiosbank_nama_biller_raw": "",
            "kiosbank_biaya_admin": None,
            "kiosbank_fee_mitra": None,
            "alterra_sheet": "",
            "alterra_product_id": "",
            "alterra_nama_produk_raw": "",
            "alterra_price": None,
            "alterra_commission": None,
            "alterra_service_fee": None,
            "alterra_collection_fee_for_partner": None,
            "catatan": "",
            "_has_alterra": False,
            "_has_kiosbank": False,
        }
        rows[sheet_name].append(row)
        by_sku[internal_sku_code] = row
        return row

    for record in alterra_records:
        sheet_name = source_sheet_for_type(record["type"])
        row = ensure_row(
            sheet_name,
            record["internal_sku_code"],
            record["category"],
            record["brand"],
            record["internal_name"],
        )
        row["alterra_sheet"] = record["sheet"]
        row["alterra_product_id"] = record["alterra_product_id"]
        row["alterra_nama_produk_raw"] = record["raw_name"]
        if record["type"] == "prepaid":
            row["alterra_price"] = record["price"]
            row["alterra_commission"] = record["commission"]
        else:
            row["alterra_service_fee"] = record["admin"]
            row["alterra_collection_fee_for_partner"] = record["commission"]
        row["_has_alterra"] = True

    for record in kiosbank_records:
        matched_row = by_sku.get(record["internal_sku_code"])
        if matched_row is not None:
            row = matched_row
            if not row["jenis_produk"]:
                row["jenis_produk"] = record["category"]
            if not row["brand"]:
                row["brand"] = record["brand"]
            if not row["nama_sku_kita"]:
                row["nama_sku_kita"] = record["raw_name"]
        else:
            row = ensure_row(
                source_sheet_for_type(record["type"]),
                record["internal_sku_code"],
                record["category"],
                record["brand"],
                record["raw_name"],
            )

        row["kiosbank_sheet"] = record["sheet"]
        row["kiosbank_kode_produk"] = record["kb_code"]
        if row["sheet_name"] == "Prepaid":
            row["kiosbank_nama_produk_raw"] = record["raw_name"]
            row["kiosbank_harga_jual"] = record["price"] or None
            row["kiosbank_fee_mitra_raw"] = record["fee_raw"]
            row["kiosbank_fee_mitra_numeric"] = record["commission"]
            if record["admin"] > 0 and record["price"] == 0:
                row["catatan"] = note_join(
                    row["catatan"],
                    record.get("note", ""),
                    f"kiosbank biaya_admin={record['admin']}",
                    f"kiosbank fee_mitra_numeric={record['commission']}",
                )
            else:
                row["catatan"] = note_join(row["catatan"], record.get("note", ""))
        else:
            row["kiosbank_nama_biller_raw"] = record["raw_name"]
            row["kiosbank_biaya_admin"] = record["admin"]
            row["kiosbank_fee_mitra"] = record["commission"]
            row["catatan"] = note_join(row["catatan"], record.get("note", ""))
        row["_has_kiosbank"] = True

    for sheet_name, sheet_rows in rows.items():
        for row in sheet_rows:
            if row["_has_alterra"] and row["_has_kiosbank"]:
                row["match_status"] = "both"
            elif row["_has_alterra"]:
                row["match_status"] = "alterra_only"
                row["catatan"] = note_join(row["catatan"], "provider-only product")
            else:
                row["match_status"] = "kiosbank_only"
                row["catatan"] = note_join(row["catatan"], "provider-only product")

        sheet_rows.sort(
            key=lambda item: (
                category_sort_rank(sheet_name, item["jenis_produk"]),
                safe_string(item["brand"]).lower(),
                row_sort_amount(sheet_name, item),
                safe_string(item["nama_sku_kita"]).lower(),
                safe_string(item["kode_produk_kita"]).lower(),
            )
        )

    return rows


def validate_rows(rows: dict[str, list[dict]], alterra_records: list[dict], kiosbank_records: list[dict]):
    issues = []
    for sheet_name, sheet_rows in rows.items():
        for row in sheet_rows:
            if not safe_string(row["jenis_produk"]):
                issues.append(f"{sheet_name}: missing jenis_produk for {row['kode_produk_kita']}")
            if not safe_string(row["brand"]):
                issues.append(f"{sheet_name}: missing brand for {row['kode_produk_kita']}")
            if not row["kiosbank_kode_produk"] and not row["alterra_product_id"]:
                issues.append(f"{sheet_name}: missing both provider codes for {row['kode_produk_kita']}")

    alterra_count = sum(1 for sheet_rows in rows.values() for row in sheet_rows if row["alterra_product_id"])
    kiosbank_count = sum(1 for sheet_rows in rows.values() for row in sheet_rows if row["kiosbank_kode_produk"])

    if alterra_count != len(alterra_records):
        issues.append(f"alterra row count mismatch: workbook={alterra_count}, parsed={len(alterra_records)}")
    if kiosbank_count != len(kiosbank_records):
        issues.append(f"kiosbank row count mismatch: workbook={kiosbank_count}, parsed={len(kiosbank_records)}")

    prepaid_telkomsel = any(
        row["match_status"] == "both"
        and row["brand"] == "TELKOMSEL"
        and "5.000" in safe_string(row["nama_sku_kita"])
        for row in rows["Prepaid"]
    )
    if not prepaid_telkomsel:
        issues.append("spot check failed: Pulsa Telkomsel 5.000 not found as merged prepaid row")

    bpjs_both = any(
        row["match_status"] == "both"
        and row["brand"] == "BPJS"
        and "BPJS" in safe_string(row["nama_sku_kita"]).upper()
        for row in rows["Postpaid"]
    )
    if not bpjs_both:
        issues.append("spot check failed: BPJS Kesehatan not found as merged postpaid row")

    markup_present = any(
        "Markup By Mitra" in safe_string(row["kiosbank_fee_mitra_raw"])
        for row in rows["Prepaid"]
    )
    if not markup_present:
        issues.append("spot check failed: Markup By Mitra not preserved in prepaid raw fee column")

    if issues:
        raise ValueError("Validation failed:\n- " + "\n- ".join(issues))


def style_sheet(ws, columns: list[str]):
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(bold=True)
    wrap = Alignment(vertical="top", wrap_text=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for idx, header in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = wrap

        max_len = len(header)
        for row in ws.iter_rows(min_row=2, min_col=idx, max_col=idx):
            value = row[0].value
            if value is None:
                continue
            max_len = max(max_len, len(str(value)))
            row[0].alignment = wrap

        ws.column_dimensions[get_column_letter(idx)].width = min(max(max_len + 2, 14), 40)


def write_sheet(ws, columns: list[str], rows: list[dict]):
    ws.append(columns)
    for row in rows:
        ws.append([row.get(column) for column in columns])
    style_sheet(ws, columns)


def write_workbook(output_path: Path, rows: dict[str, list[dict]]):
    workbook = openpyxl.Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    prepaid_ws = workbook.create_sheet("Prepaid")
    write_sheet(prepaid_ws, PREPAID_COLUMNS, rows["Prepaid"])

    postpaid_ws = workbook.create_sheet("Postpaid")
    write_sheet(postpaid_ws, POSTPAID_COLUMNS, rows["Postpaid"])

    output_path.parent.mkdir(parents=True, exist_ok=True)
    try:
        workbook.save(output_path)
        return output_path
    except PermissionError:
        timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
        fallback = output_path.with_name(f"{output_path.stem} {timestamp}{output_path.suffix}")
        workbook.save(fallback)
        return fallback


def summarize(rows: dict[str, list[dict]]):
    for sheet_name in ("Prepaid", "Postpaid"):
        sheet_rows = rows[sheet_name]
        both = sum(1 for row in sheet_rows if row["match_status"] == "both")
        alterra_only = sum(1 for row in sheet_rows if row["match_status"] == "alterra_only")
        kiosbank_only = sum(1 for row in sheet_rows if row["match_status"] == "kiosbank_only")
        print(f"{sheet_name}: total={len(sheet_rows)}, both={both}, alterra_only={alterra_only}, kiosbank_only={kiosbank_only}")


def parse_args():
    parser = argparse.ArgumentParser(description="Build docs-only provider catalog workbook")
    parser.add_argument("--output", default=str(DEFAULT_OUTPUT), help="Output workbook path")
    return parser.parse_args()


def main():
    args = parse_args()
    output_path = Path(args.output)

    alterra_records = build_alterra_records()
    alterra_lookup = build_alterra_lookup(alterra_records)
    kiosbank_records = build_kiosbank_records()
    assign_kiosbank_internal_codes(kiosbank_records, alterra_lookup)
    rows = build_rows(alterra_records, kiosbank_records)
    validate_rows(rows, alterra_records, kiosbank_records)
    written_path = write_workbook(output_path, rows)

    print(f"Alterra parsed: {len(alterra_records)}")
    print(f"Kiosbank parsed: {len(kiosbank_records)}")
    summarize(rows)
    print(f"Workbook written: {written_path}")


if __name__ == "__main__":
    main()
