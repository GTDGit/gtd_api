"""
Parse Kiosbank pricelist Excel and generate SQL migration.

Usage:
  py scripts/parse_kiosbank_pricelist.py

Reads: docs/ppob/kiosbank/Pricelist Kiosbank.xlsx
Writes: migrations/000029_seed_kiosbank_products.up.sql
        migrations/000029_seed_kiosbank_products.down.sql

Pricing model:
  Postpaid: admin = Biaya Administrasi (our cost), commission = Fee Mitra (our revenue)
  Prepaid:  price = Harga Jual (buy price), commission = Fee Mitra if numeric (else 0)
  Markup:   products.admin = ~0.2% of price (prepaid), or suggested admin (postpaid)

Product matching:
  Products that exist in Alterra get Kiosbank SKU linked to SAME product row.
  Kiosbank-only products get new product with KB- prefix.
"""

import openpyxl
import os
import re

EXCEL_PATH = os.path.join(os.path.dirname(__file__), "..", "docs", "ppob", "kiosbank",
                          "Pricelist Kiosbank.xlsx")
ALTERRA_SQL = os.path.join(os.path.dirname(__file__), "..", "migrations",
                           "000028_reseed_alterra_pricing.up.sql")
UP_PATH = os.path.join(os.path.dirname(__file__), "..", "migrations",
                       "000029_seed_kiosbank_products.up.sql")
DOWN_PATH = os.path.join(os.path.dirname(__file__), "..", "migrations",
                         "000029_seed_kiosbank_products.down.sql")


def escape_sql(s):
    if s is None:
        return ""
    return str(s).replace("'", "''").strip()


def safe_int(val):
    if val is None:
        return 0
    if isinstance(val, (int, float)):
        return int(val)
    s = str(val).strip().replace(",", "").replace(".", "")
    if s == "-" or s == "":
        return 0
    try:
        return int(s)
    except (ValueError, TypeError):
        return 0


def parse_fee(val):
    """Parse Fee Mitra: can be int, 'Markup By Mitra', '-', or None."""
    if val is None:
        return 0
    if isinstance(val, (int, float)):
        return int(val)
    s = str(val).strip()
    if s.lower().startswith("markup") or s == "-" or s == "":
        return 0
    try:
        return int(s.replace(",", "").replace(".", ""))
    except (ValueError, TypeError):
        return 0


def calc_markup(price_or_denom):
    if price_or_denom <= 0:
        return 0
    markup = round(price_or_denom * 0.002 / 10) * 10
    return max(10, min(2000, markup))


def normalize_brand(name):
    """Normalize brand/operator name to match Alterra's conventions."""
    if not name:
        return "LAINNYA"
    n = str(name).strip().lower()

    # Mobile operators
    if "telkomsel" in n or n == "tsel":
        return "TELKOMSEL"
    if ("xl" in n and "axis" in n) or n == "xl & axis":
        return "XL"
    if n.startswith("xl") and "axis" not in n:
        return "XL"
    if "axis" in n:
        return "AXIS"
    if n in ("three", "tri") or n.startswith("three ") or n.startswith("tri "):
        return "TRI"
    if "smartfren" in n or "smart/fren" in n:
        return "SMARTFREN"
    if "indosat" in n or "im3" in n:
        return "INDOSAT"
    if "by.u" in n or "byu" in n:
        return "BYU"

    # Telco
    if "telkom" in n and "sel" not in n and "halo" not in n:
        return "TELKOM"
    if "indihome" in n:
        return "INDIHOME"
    if "halo" in n:
        return "TELKOMSEL"

    # E-wallet
    if "dana" in n:
        return "DANA"
    if "gopay" in n:
        return "GOPAY"
    if "ovo" in n:
        return "OVO"
    if "shopee" in n:
        return "SHOPEEPAY"
    if "brizzi" in n:
        return "BRIZZI"
    if "flazz" in n:
        return "FLAZZ"
    if "tapcash" in n:
        return "BNI TAPCASH"
    if "mandiri" in n and ("e-money" in n or "emoney" in n or "e money" in n):
        return "MANDIRI E-MONEY"
    if "linkaja" in n:
        return "LINKAJA"

    # TV
    if "mnc" in n:
        return "MNC VISION"
    if "transvision" in n:
        return "TRANSVISION"
    if "k-vision" in n or "kvision" in n:
        return "K-VISION"
    if "nex" in n and "parabola" in n:
        return "NEX PARABOLA"
    if "cbn" in n:
        return "CBN"

    # Internet
    if "myrepublic" in n or "my republic" in n:
        return "MY REPUBLIC"
    if "firstmedia" in n or "first media" in n:
        return "FIRST MEDIA"
    if "oxygen" in n:
        return "OXYGEN"
    if "globalxtreme" in n:
        return "GLOBALXTREME"

    # Gas
    if "pgn" in n:
        return "PGN"

    # Clean up
    brand = str(name).strip().upper()
    brand = re.sub(r'[^A-Z0-9 &\-]', '', brand)
    return brand if brand else "LAINNYA"


def extract_denom(name):
    """Extract denomination from product name like 'Telkomsel Prabayar 50.000' -> 50000."""
    if not name:
        return 0
    m = re.search(r'([\d.]+)\s*$', str(name).strip())
    if m:
        val = m.group(1).replace(".", "")
        try:
            return int(val)
        except ValueError:
            return 0
    return 0


def is_data_row(row):
    """Check if row is a product data row (has product code in col[1])."""
    if row[1] is None:
        return False
    code = str(row[1]).strip()
    if code == "" or code.lower() in ("none",):
        return False
    return True


# ============================================================
# Build Alterra product lookup from migration 000028
# ============================================================

def build_alterra_lookup():
    """Parse migration 000028 SQL to extract existing products for matching."""
    lookup = {}  # key: (category_lower, brand_lower, denom_or_name) -> sku_code

    if not os.path.exists(ALTERRA_SQL):
        print(f"Warning: {ALTERRA_SQL} not found, no product matching")
        return lookup

    with open(ALTERRA_SQL, "r", encoding="utf-8") as f:
        content = f.read()

    # Parse INSERT INTO products lines
    pattern = re.compile(
        r"INSERT INTO products.*?VALUES\s*\("
        r"'([^']+)',\s*"  # sku_code
        r"'([^']*)',\s*"  # name
        r"'([^']*)',\s*"  # category
        r"'([^']*)',\s*"  # brand
        r"'([^']*)',\s*"  # type
        r"(\d+)"          # admin
    )

    for m in pattern.finditer(content):
        sku_code = m.group(1)
        name = m.group(2)
        category = m.group(3)
        brand = m.group(4)
        ptype = m.group(5)

        cat_lower = category.lower()
        brand_lower = brand.lower()
        name_lower = name.lower()

        # Key 1: category + brand + type + denomination (most specific)
        denom = extract_denom(name)
        if denom > 0:
            key = (cat_lower, brand_lower, ptype, denom)
            lookup[key] = sku_code

        # Key 2: category + brand + type (for single-product postpaid categories)
        key2 = (cat_lower, brand_lower, ptype, 0)
        if key2 not in lookup:
            lookup[key2] = sku_code

        # Key 3: exact name match
        lookup[("name", name_lower, 0, 0)] = sku_code

    print(f"Built Alterra lookup: {len(lookup)} entries")
    return lookup


def find_matching_product(lookup, category, brand, ptype, denom, name=""):
    """Try to find matching Alterra product."""
    cat = category.lower()
    br = brand.lower()

    # Try category + brand + type + denom (most specific)
    if denom > 0:
        key = (cat, br, ptype, denom)
        if key in lookup:
            return lookup[key]

    # Try category + brand + type (for postpaid single-entry products)
    key2 = (cat, br, ptype, 0)
    if key2 in lookup:
        return lookup[key2]

    # Try without type constraint (denom match)
    if denom > 0:
        for t in ("prepaid", "postpaid"):
            key3 = (cat, br, t, denom)
            if key3 in lookup:
                return lookup[key3]

    # Try name match
    if name:
        key4 = ("name", name.lower(), 0, 0)
        if key4 in lookup:
            return lookup[key4]

    return None


# ============================================================
# Sheet parsers
# ============================================================

def parse_postpaid_sheet(ws, category, brand_from_name=False, fixed_brand=None):
    """Generic parser for postpaid sheets with columns:
    No, Kode Produk, Nama Billing Provider, Biaya Administrasi, Fee Mitra"""
    products = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not is_data_row(row):
            continue

        code = str(row[1]).strip()
        name = str(row[2]).strip() if row[2] else code
        admin = safe_int(row[3])
        commission = parse_fee(row[4])

        if fixed_brand:
            brand = fixed_brand
        elif brand_from_name:
            brand = normalize_brand(name)
        else:
            brand = normalize_brand(category)

        products.append({
            "kb_code": code,
            "name": name,
            "category": category,
            "brand": brand,
            "type": "postpaid",
            "price": 0,
            "admin": admin,
            "commission": commission,
            "suggested_admin": admin,  # customer admin = Biaya Administrasi from Kiosbank
        })
    return products


def parse_pln(ws):
    """PLN sheet: mix of prepaid tokens and postpaid billing."""
    products = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not is_data_row(row):
            continue

        code = str(row[1]).strip()
        name = str(row[2]).strip() if row[2] else code
        admin = safe_int(row[3])
        commission = parse_fee(row[4])

        name_lower = name.lower()
        if "prepaid" in name_lower:
            # PLN token - prepaid but uses admin/fee structure
            denom = extract_denom(name)
            products.append({
                "kb_code": code,
                "name": name,
                "category": "Listrik",
                "brand": "PLN",
                "type": "prepaid",
                "price": 0,
                "admin": admin,
                "commission": commission,
                "suggested_admin": admin,
            })
        else:
            # PLN Postpaid or Non Taglis
            products.append({
                "kb_code": code,
                "name": name,
                "category": "Listrik",
                "brand": "PLN",
                "type": "postpaid",
                "price": 0,
                "admin": admin,
                "commission": commission,
                "suggested_admin": admin,
            })
    return products


def parse_pgn(ws):
    """PGN sheet: mix of postpaid and prepaid."""
    products = []
    section = "postpaid"
    for row in ws.iter_rows(min_row=2, values_only=True):
        # Check section headers
        val0 = str(row[0]).strip().upper() if row[0] else ""
        if "PREPAID" in val0:
            section = "prepaid"
            continue
        if "POSTPAID" in val0:
            section = "postpaid"
            continue

        if not is_data_row(row):
            continue

        code = str(row[1]).strip()
        name = str(row[2]).strip() if row[2] else code
        admin = safe_int(row[3])
        commission = parse_fee(row[4])

        products.append({
            "kb_code": code,
            "name": name,
            "category": "Gas PGN",
            "brand": "PGN",
            "type": section,
            "price": 0,
            "admin": admin,
            "commission": commission,
            "suggested_admin": admin,
        })
    return products


def parse_prepaid_sheet(ws, category, brand_from_sections=False, fixed_brand=None,
                        price_col=3, fee_col=4, name_col=2, code_col=1):
    """Generic parser for prepaid sheets with columns:
    No, Kode Produk, Nama Produk, Harga Jual, Fee Mitra"""
    products = []
    current_brand = fixed_brand or "LAINNYA"

    for row in ws.iter_rows(min_row=1, values_only=True):
        # Detect brand section headers (text in col0, nothing in col1)
        if brand_from_sections and row[code_col] is None and row[0] is not None:
            header = str(row[0]).strip()
            if header and not header[0].isdigit() and not header.lower().startswith(("no", "ketentuan", "1.", "2.")):
                current_brand = normalize_brand(header)
                continue

        if row[code_col] is None:
            continue

        code = str(row[code_col]).strip()
        # Skip header rows
        if code.lower() in ("kode produk", "id product", "no.", "no"):
            continue

        name = str(row[name_col]).strip() if row[name_col] else code
        price = safe_int(row[price_col])
        commission = parse_fee(row[fee_col])

        if price <= 0:
            continue

        if not brand_from_sections and not fixed_brand:
            # Extract brand from first word(s) of product name
            brand = normalize_brand(name.split(" ")[0] if " " in name else name)
        else:
            brand = current_brand

        products.append({
            "kb_code": code,
            "name": name,
            "category": category,
            "brand": brand,
            "type": "prepaid",
            "price": price,
            "admin": 0,
            "commission": commission,
            "suggested_admin": calc_markup(price),
        })
    return products


def parse_uang_elektronik(ws):
    """UANG ELEKTRONIK: mixed format.
    First section (DANA, ShopeePay, GoPay): Id Product, Product, Harga Jual, Fee Mitra
    Second section (OVO etc): different header row mid-sheet."""
    products = []
    for row in ws.iter_rows(min_row=1, values_only=True):
        # col0=Id Product, col1=Product name, col2=Harga Jual, col3=Fee Mitra
        if row[0] is None:
            continue

        code_val = row[0]
        # Skip header rows
        if str(code_val).strip().lower() in ("id product", "biller", ""):
            continue

        # Check if this is a product ID (numeric or alphanumeric code)
        code = str(code_val).strip()
        try:
            int(code)
        except ValueError:
            continue

        name = str(row[1]).strip() if row[1] else code
        price = safe_int(row[2])
        commission = parse_fee(row[3])

        if price <= 0:
            continue

        brand = normalize_brand(name)

        products.append({
            "kb_code": code,
            "name": name,
            "category": "E-Money",
            "brand": brand,
            "type": "prepaid",
            "price": price,
            "admin": 0,
            "commission": commission,
            "suggested_admin": calc_markup(price),
        })
    return products


def parse_nexparabola(ws):
    """NEXPARABOLA: No, Id Product, Biller, Harga Jual, Fee Mitra."""
    products = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not is_data_row(row):
            continue

        code = str(row[1]).strip()
        name = str(row[2]).strip() if row[2] else code
        price = safe_int(row[3])
        commission = parse_fee(row[4])

        if price <= 0:
            continue

        products.append({
            "kb_code": code,
            "name": name,
            "category": "TV Kabel",
            "brand": "NEX PARABOLA",
            "type": "prepaid",
            "price": price,
            "admin": 0,
            "commission": commission,
            "suggested_admin": calc_markup(price),
        })
    return products


# ============================================================
# Category mapping for matching with Alterra
# ============================================================

CATEGORY_MAP = {
    "Pulsa": "Pulsa",
    "Paket Data": "Paket Data",
    "Listrik": "Listrik",
    "BPJS Kesehatan": "BPJS Kesehatan",
    "PDAM": "PDAM",
    "PBB": "PBB",
    "Telepon Pascabayar": "Telepon Pascabayar",
    "Internet": "Internet",
    "TV Kabel": "TV Kabel",
    "Gas PGN": "Gas PGN",
    "E-Money": "E-Money",
    "Multifinance": "Multifinance",
    "Voucher": "Voucher",
    "Game": "Voucher Game",
}


# ============================================================
# SQL generation
# ============================================================

def generate_sql(all_products, alterra_lookup):
    """Generate SQL migration."""
    # Deduplicate by kb_code
    seen = set()
    unique = []
    for p in all_products:
        if p["kb_code"] not in seen:
            seen.add(p["kb_code"])
            unique.append(p)

    # Categories where each product is unique (per-region billing providers)
    # These should NOT match to a single Alterra generic product
    NO_MATCH_CATEGORIES = {"PDAM", "PBB", "Multifinance"}

    # Try to match each product to existing Alterra product
    matched = 0
    new_products = 0
    for p in unique:
        cat = p["category"]
        brand = p["brand"]
        denom = extract_denom(p["name"])

        if cat in NO_MATCH_CATEGORIES:
            sku = None  # Always create new product for these
        else:
            sku = find_matching_product(alterra_lookup, cat, brand, p["type"], denom, p["name"])
        if sku:
            p["product_sku"] = sku
            p["is_new"] = False
            matched += 1
        else:
            # Generate KB- sku_code for new product
            cat_slug = re.sub(r'[^A-Z0-9]', '', cat.upper())[:10]
            brand_slug = re.sub(r'[^A-Z0-9]', '', brand.upper())[:10]
            p["product_sku"] = f"KB-{cat_slug}-{brand_slug}-{p['kb_code']}"
            p["is_new"] = True
            new_products += 1

    # Deduplicate product_sku
    sku_seen = {}
    for p in unique:
        sku = p["product_sku"]
        if sku in sku_seen:
            p["product_sku"] = f"{sku}-{p['kb_code']}"
        sku_seen[p["product_sku"]] = True

    print(f"Matched: {matched}, New: {new_products}, Total: {len(unique)}")

    # Collect unique new categories and brands
    new_categories = set()
    new_brands = set()
    for p in unique:
        new_categories.add(p["category"])
        new_brands.add(p["brand"])

    lines = []
    lines.append("-- ============================================")
    lines.append("-- Migration 000029: Seed Kiosbank Products")
    lines.append("-- Auto-generated by scripts/parse_kiosbank_pricelist.py")
    lines.append("-- Kiosbank pricing: postpaid admin/commission, prepaid price")
    lines.append("-- Matched products link to existing Alterra products")
    lines.append("-- ============================================")
    lines.append("")
    lines.append("BEGIN;")
    lines.append("")

    # 1. Clean existing Kiosbank provider SKUs
    lines.append("-- 1. Clean Kiosbank provider SKUs")
    lines.append("DELETE FROM ppob_provider_skus WHERE provider_id = (SELECT id FROM ppob_providers WHERE code = 'kiosbank');")
    lines.append("")

    # 2. Add new categories
    lines.append("-- 2. Seed categories")
    cat_values = ", ".join(f"('{escape_sql(c)}')" for c in sorted(new_categories))
    lines.append(f"INSERT INTO product_categories (name) VALUES {cat_values}")
    lines.append("ON CONFLICT (name) DO NOTHING;")
    lines.append("")

    # 3. Add new brands
    lines.append("-- 3. Seed brands")
    brand_values = ", ".join(f"('{escape_sql(b)}')" for b in sorted(new_brands))
    lines.append(f"INSERT INTO product_brands (name) VALUES {brand_values}")
    lines.append("ON CONFLICT (name) DO NOTHING;")
    lines.append("")

    # 4. Upsert new products (KB- prefix only)
    new_prods = [p for p in unique if p["is_new"]]
    if new_prods:
        lines.append(f"-- 4. Upsert new Kiosbank-only products ({len(new_prods)} products)")
        for p in new_prods:
            product_admin = p["suggested_admin"]
            lines.append(
                f"INSERT INTO products (sku_code, name, category, brand, type, admin, commission, description, is_active) "
                f"VALUES ('{escape_sql(p['product_sku'])}', '{escape_sql(p['name'])}', '{escape_sql(p['category'])}', "
                f"'{escape_sql(p['brand'])}', '{p['type']}', {product_admin}, 0, "
                f"'{escape_sql(p['name'])}', true) "
                f"ON CONFLICT (sku_code) DO UPDATE SET name = EXCLUDED.name, category = EXCLUDED.category, "
                f"brand = EXCLUDED.brand, type = EXCLUDED.type, admin = EXCLUDED.admin, "
                f"description = EXCLUDED.description;"
            )
        lines.append("")

    # 5. Insert Kiosbank provider SKU mappings
    lines.append(f"-- 5. Insert Kiosbank provider SKU mappings ({len(unique)} SKUs)")
    for p in unique:
        price_val = p["price"]
        admin_val = p["admin"]
        commission_val = p["commission"]

        lines.append(
            f"INSERT INTO ppob_provider_skus (provider_id, product_id, provider_sku_code, provider_product_name, price, admin, commission, is_active, is_available) "
            f"SELECT (SELECT id FROM ppob_providers WHERE code = 'kiosbank'), p.id, '{escape_sql(p['kb_code'])}', '{escape_sql(p['name'])}', "
            f"{price_val}, {admin_val}, {commission_val}, true, true "
            f"FROM products p WHERE p.sku_code = '{escape_sql(p['product_sku'])}' "
            f"ON CONFLICT (provider_id, provider_sku_code) DO UPDATE SET "
            f"price = EXCLUDED.price, admin = EXCLUDED.admin, commission = EXCLUDED.commission, "
            f"provider_product_name = EXCLUDED.provider_product_name, "
            f"is_active = true, is_available = true;"
        )

    lines.append("")
    lines.append("-- 6. Ensure Kiosbank provider is active")
    lines.append("UPDATE ppob_providers SET is_active = true WHERE code = 'kiosbank';")
    lines.append("")
    lines.append("COMMIT;")

    return "\n".join(lines)


def generate_down_sql():
    return (
        "-- Rollback: remove Kiosbank provider SKUs\n"
        "DELETE FROM ppob_provider_skus WHERE provider_id = (SELECT id FROM ppob_providers WHERE code = 'kiosbank');\n"
        "\n"
        "-- Note: New KB- products are not removed (may have FK refs from transactions).\n"
    )


def main():
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True)
    alterra_lookup = build_alterra_lookup()

    all_products = []

    # Postpaid sheets
    postpaid_sheets = [
        ("PDAM", "PDAM", False, "PDAM"),
        ("PBB", "PBB", False, "PBB"),
        ("MULTIFINANCE", "Multifinance", True, None),
        ("BPJS", "BPJS Kesehatan", False, "BPJS"),
        ("INTERNET&TV KABEL", "Internet", True, None),
        ("TELEKOMUNIKASI", "Telepon Pascabayar", True, None),
        ("TV BERBAYAR", "TV Kabel", True, None),
    ]

    for sheet_name, category, brand_from_name, fixed_brand in postpaid_sheets:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            products = parse_postpaid_sheet(ws, category, brand_from_name, fixed_brand)
            print(f"  {sheet_name}: {len(products)} products")
            all_products.extend(products)

    # PLN (mixed)
    if "PLN" in wb.sheetnames:
        products = parse_pln(wb["PLN"])
        print(f"  PLN: {len(products)} products")
        all_products.extend(products)

    # PGN (mixed)
    if "PGN" in wb.sheetnames:
        products = parse_pgn(wb["PGN"])
        print(f"  PGN: {len(products)} products")
        all_products.extend(products)

    # Prepaid sheets with brand sections
    if "PULSA" in wb.sheetnames:
        products = parse_prepaid_sheet(wb["PULSA"], "Pulsa", brand_from_sections=True)
        print(f"  PULSA: {len(products)} products")
        all_products.extend(products)

    if "PAKET DATA" in wb.sheetnames:
        products = parse_prepaid_sheet(wb["PAKET DATA"], "Paket Data", brand_from_sections=True)
        print(f"  PAKET DATA: {len(products)} products")
        all_products.extend(products)

    if "VOUCHER GAME" in wb.sheetnames:
        products = parse_prepaid_sheet(wb["VOUCHER GAME"], "Voucher Game", brand_from_sections=True)
        print(f"  VOUCHER GAME: {len(products)} products")
        all_products.extend(products)

    # Prepaid sheets with brand from name
    if "VOUCHER BELANJA" in wb.sheetnames:
        products = parse_prepaid_sheet(wb["VOUCHER BELANJA"], "Voucher", brand_from_sections=False)
        print(f"  VOUCHER BELANJA: {len(products)} products")
        all_products.extend(products)

    # Nex Parabola
    if "NEXPARABOLA" in wb.sheetnames:
        products = parse_nexparabola(wb["NEXPARABOLA"])
        print(f"  NEXPARABOLA: {len(products)} products")
        all_products.extend(products)

    # Uang Elektronik
    if "UANG ELEKTRONIK" in wb.sheetnames:
        products = parse_uang_elektronik(wb["UANG ELEKTRONIK"])
        print(f"  UANG ELEKTRONIK: {len(products)} products")
        all_products.extend(products)

    wb.close()

    print(f"\nTotal products parsed: {len(all_products)}")

    # Generate SQL
    sql = generate_sql(all_products, alterra_lookup)

    with open(UP_PATH, "w", encoding="utf-8") as f:
        f.write(sql)
    print(f"Written: {UP_PATH} ({len(sql.splitlines())} lines)")

    down_sql = generate_down_sql()
    with open(DOWN_PATH, "w", encoding="utf-8") as f:
        f.write(down_sql)
    print(f"Written: {DOWN_PATH}")

    # Print summary
    print("\n=== SUMMARY ===")
    cats = {}
    for p in all_products:
        key = (p["category"], p["type"])
        cats[key] = cats.get(key, 0) + 1
    for (cat, ptype), count in sorted(cats.items()):
        print(f"  {cat} ({ptype}): {count}")


if __name__ == "__main__":
    main()
