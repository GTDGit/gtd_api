"""
Parse Alterra product catalog Excel and generate SQL migration.

Usage:
  python scripts/parse_alterra_catalog.py

Reads: docs/ppob/alterra/20262801 - BPA Product Catalogue  (OPEN).xlsx
Writes: migrations/000028_reseed_alterra_pricing.up.sql
        migrations/000028_reseed_alterra_pricing.down.sql

Pricing model:
  Prepaid:  price = buy price (service_fee or sell_price), commission = collection_fee if any
  Postpaid: admin = service_fee (our cost per trx), commission = collection_fee_for_partner (our revenue)
  Markup:   products.admin = ~0.2% of price/denom, min 10, max 2000 (prepaid)
            products.admin = suggested admin from Excel (postpaid)
"""

import openpyxl
import os
import re

EXCEL_PATH = os.path.join(os.path.dirname(__file__), "..", "docs", "ppob", "alterra",
                          "20262801 - BPA Product Catalogue  (OPEN).xlsx")
UP_PATH = os.path.join(os.path.dirname(__file__), "..", "migrations", "000028_reseed_alterra_pricing.up.sql")
DOWN_PATH = os.path.join(os.path.dirname(__file__), "..", "migrations", "000028_reseed_alterra_pricing.down.sql")


def escape_sql(s):
    if s is None:
        return ""
    return str(s).replace("'", "''").strip()


def format_rupiah(amount):
    """Format number as Indonesian Rupiah string: 5000 -> '5.000'"""
    s = str(int(amount))
    result = []
    for i, c in enumerate(reversed(s)):
        if i > 0 and i % 3 == 0:
            result.append(".")
        result.append(c)
    return "".join(reversed(result))


def normalize_brand(operator):
    """Normalize operator name to standard brand."""
    if not operator:
        return "LAINNYA"
    op = str(operator).strip().lower()

    # Mobile operators
    if "axis" in op:
        return "AXIS"
    if "xl" in op and "axis" not in op:
        return "XL"
    if "telkomsel" in op:
        return "TELKOMSEL"
    if "indosat" in op or "im3" in op or "ooredoo" in op:
        return "INDOSAT"
    if "tri" in op or "three" in op:
        return "TRI"
    if "smartfren" in op:
        return "SMARTFREN"
    if "by.u" in op or "byu" in op:
        return "BYU"

    # Telco
    if op == "pln":
        return "PLN"
    if "telkom" in op:
        return "TELKOM"

    # E-wallet brands
    if "dana" in op:
        return "DANA"
    if "gopay" in op:
        return "GOPAY"
    if "ovo" in op:
        return "OVO"
    if "shopee" in op:
        return "SHOPEEPAY"
    if "linkaja" in op:
        return "LINKAJA"
    if "mandiri" in op or "e-money" in op:
        return "MANDIRI E-MONEY"
    if "bni" in op or "tapcash" in op:
        return "BNI TAPCASH"
    if "brizzi" in op or "bri" in op:
        return "BRIZZI"
    if "maxim" in op:
        return "MAXIM"
    if "grab" in op:
        return "GRAB"

    # Clean up generic
    brand = str(operator).strip().upper()
    brand = re.sub(r'[^A-Z0-9 &\-]', '', brand)
    return brand if brand else "LAINNYA"


def safe_int(val):
    """Safely convert value to int."""
    if val is None:
        return 0
    if isinstance(val, (int, float)):
        return int(val)
    try:
        cleaned = str(val).replace(",", "").replace(".", "").strip()
        return int(cleaned) if cleaned else 0
    except (ValueError, TypeError):
        return 0


def safe_float(val):
    """Safely convert value to float (for percentage commission)."""
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    try:
        return float(str(val).strip())
    except (ValueError, TypeError):
        return 0.0


def is_product_id(val):
    """Check if value looks like a product ID (integer)."""
    return isinstance(val, (int, float)) and val is not None and val > 0


def calc_markup(price_or_denom):
    """Calculate markup proportional to denomination.
    ~0.2% of price, min 10, max 2000, rounded to nearest 10."""
    if price_or_denom <= 0:
        return 0
    markup = round(price_or_denom * 0.002 / 10) * 10
    return max(10, min(2000, markup))


def parse_suggested_admin(val):
    """Parse suggested admin fee from Excel (handles strings like '2,750 - 3,000')."""
    if val is None:
        return 0
    if isinstance(val, (int, float)):
        return int(val)
    s = str(val).strip()
    # Try to extract first number from range like "2,750 - 3,000"
    m = re.search(r'[\d,]+', s)
    if m:
        return safe_int(m.group(0).replace(",", ""))
    return 0


# ============================================================
# Sheet parsers - each returns list of product dicts with:
#   sku_code, name, category, brand, type, description,
#   alterra_product_id, price, admin, commission, suggested_admin
# ============================================================

def parse_pulsa(ws):
    """Parse Pulsa sheet - prepaid mobile top-up.
    Cols: Product ID, Tipe, Produk, Denom, Operator, Price Type, Sell Price"""
    products = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        vals = list(row)[:7]
        product_id = vals[0]
        if not is_product_id(product_id):
            continue

        product_id = int(product_id)
        name = str(vals[2]).strip() if vals[2] else ""
        denom = safe_int(vals[3])
        operator = str(vals[4]).strip() if vals[4] else ""
        sell_price = safe_int(vals[6])

        brand = normalize_brand(operator)
        display_name = f"Rp {format_rupiah(denom)}" if denom > 0 else name
        sku_code = f"ALT-PULSA-{brand}-{denom}" if denom > 0 else f"ALT-PULSA-{brand}-{product_id}"

        products.append({
            "sku_code": sku_code, "name": display_name,
            "category": "Pulsa", "brand": brand, "type": "prepaid",
            "description": name, "alterra_product_id": product_id,
            "price": sell_price, "admin": 0, "commission": 0,
            "suggested_admin": calc_markup(sell_price),
        })
    return products


def parse_pln_bpjs_postpaid(ws):
    """Parse 'PLN, BPJS & Postpaid' sheet - multiple sections with varying column meanings.
    Key: column header row tells us whether col 4 is 'Service Fee Sepulsa' or 'Collection Fee for Partner'."""
    products = []
    current_section = None
    col4_is_collection_fee = False  # Track what column 4 means

    for row in ws.iter_rows(values_only=True):
        vals = list(row)[:8]
        first_val = vals[0]

        # Detect section headers
        if isinstance(first_val, str):
            lower = first_val.lower().strip()
            if "pln prepaid" in lower:
                current_section = "pln_prepaid"
                col4_is_collection_fee = False
                continue
            elif "pln postpaid" in lower:
                current_section = "pln_postpaid"
                col4_is_collection_fee = False
                continue
            elif "bpjs" in lower and ("product" in lower or "produk" in lower):
                current_section = "bpjs"
                col4_is_collection_fee = False
                continue
            elif "telkom postpaid" in lower:
                current_section = "telkom_postpaid"
                col4_is_collection_fee = False
                continue
            elif "mobile postpaid" in lower:
                current_section = "mobile_postpaid"
                col4_is_collection_fee = False
                continue
            elif "internet service" in lower:
                current_section = "internet_service"
                col4_is_collection_fee = False
                continue
            elif "insurance" in lower:
                current_section = "insurance"
                col4_is_collection_fee = False
                continue

            # Detect column header rows - crucial for knowing if col4 = service_fee or collection_fee
            if "product id" in lower:
                header_str = " ".join(str(v).lower() for v in vals if v)
                col4_is_collection_fee = "collection fee" in header_str
                continue
            continue

        if not is_product_id(first_val):
            continue

        product_id = int(first_val)
        tipe = str(vals[1]).strip() if vals[1] else ""
        name = str(vals[2]).strip() if vals[2] else ""
        operator = str(vals[3]).strip() if vals[3] else ""
        fee_val = safe_int(vals[4])
        suggested = parse_suggested_admin(vals[5])

        if current_section == "insurance":
            # Skip insurance products for now
            continue

        if current_section == "pln_prepaid":
            denom = 0
            m = re.search(r'[\d,]+', name.replace("PLN Prepaid Rp. ", "").replace("PLN Prepaid Rp.", ""))
            if m:
                denom = safe_int(m.group(0).replace(",", ""))
            display_name = f"Token Rp {format_rupiah(denom)}" if denom > 0 else name
            sku_code = f"ALT-LISTRIK-PLN-{denom}" if denom > 0 else f"ALT-LISTRIK-PLN-{product_id}"
            products.append({
                "sku_code": sku_code, "name": display_name,
                "category": "Listrik", "brand": "PLN", "type": "prepaid",
                "description": name, "alterra_product_id": product_id,
                "price": fee_val, "admin": 0, "commission": 0,
                "suggested_admin": calc_markup(fee_val),
            })

        elif current_section == "pln_postpaid":
            # admin = service_fee (1150), commission = 0
            products.append({
                "sku_code": f"ALT-LISTRIK-PLN-POSTPAID-{product_id}",
                "name": "Tagihan Listrik", "category": "Listrik", "brand": "PLN",
                "type": "postpaid", "description": name,
                "alterra_product_id": product_id,
                "price": 0, "admin": fee_val, "commission": 0,
                "suggested_admin": suggested if suggested > 0 else 2750,
            })

        elif current_section == "bpjs":
            if col4_is_collection_fee:
                # BPJS Ketenagakerjaan: admin=0, commission=collection_fee
                products.append({
                    "sku_code": f"ALT-BPJS-{product_id}",
                    "name": name if name else "BPJS Ketenagakerjaan",
                    "category": "BPJS Kesehatan", "brand": "BPJS", "type": "postpaid",
                    "description": name, "alterra_product_id": product_id,
                    "price": 0, "admin": 0, "commission": fee_val,
                    "suggested_admin": 2500,
                })
            else:
                # BPJS Kesehatan: admin=service_fee, commission=0
                products.append({
                    "sku_code": f"ALT-BPJS-{product_id}",
                    "name": name if name else "BPJS Kesehatan",
                    "category": "BPJS Kesehatan", "brand": "BPJS", "type": "postpaid",
                    "description": name, "alterra_product_id": product_id,
                    "price": 0, "admin": fee_val, "commission": 0,
                    "suggested_admin": suggested if suggested > 0 else 2500,
                })

        elif current_section == "telkom_postpaid":
            products.append({
                "sku_code": f"ALT-TELPASCABAYAR-TELKOM-{product_id}",
                "name": name if name else "Tagihan Telkom",
                "category": "Telepon Pascabayar", "brand": "TELKOM", "type": "postpaid",
                "description": name, "alterra_product_id": product_id,
                "price": 0, "admin": fee_val, "commission": 0,
                "suggested_admin": suggested if suggested > 0 else 2500,
            })

        elif current_section == "mobile_postpaid":
            brand = normalize_brand(operator)
            if col4_is_collection_fee:
                # XL, Indosat, Tri, Smartfren: admin=0, commission=collection_fee
                products.append({
                    "sku_code": f"ALT-TELPASCABAYAR-{brand}-{product_id}",
                    "name": name if name else f"Pascabayar {brand}",
                    "category": "Telepon Pascabayar", "brand": brand, "type": "postpaid",
                    "description": name, "alterra_product_id": product_id,
                    "price": 0, "admin": 0, "commission": fee_val,
                    "suggested_admin": 2500,
                })
            else:
                # Halo: admin=service_fee, commission=0
                products.append({
                    "sku_code": f"ALT-TELPASCABAYAR-{brand}-{product_id}",
                    "name": name if name else f"Pascabayar {brand}",
                    "category": "Telepon Pascabayar", "brand": brand, "type": "postpaid",
                    "description": name, "alterra_product_id": product_id,
                    "price": 0, "admin": fee_val, "commission": 0,
                    "suggested_admin": suggested if suggested > 0 else 2500,
                })

        elif current_section == "internet_service":
            brand = normalize_brand(operator)
            if col4_is_collection_fee:
                # Collection fee model: admin=0, commission=collection_fee
                products.append({
                    "sku_code": f"ALT-INTERNET-{brand}-{product_id}",
                    "name": name if name else f"Internet {brand}",
                    "category": "Internet", "brand": brand, "type": "postpaid",
                    "description": name, "alterra_product_id": product_id,
                    "price": 0, "admin": 0, "commission": fee_val,
                    "suggested_admin": 3500,
                })
            else:
                # Service fee model: admin=service_fee, commission=0
                products.append({
                    "sku_code": f"ALT-INTERNET-{brand}-{product_id}",
                    "name": name if name else f"Internet {brand}",
                    "category": "Internet", "brand": brand, "type": "postpaid",
                    "description": name, "alterra_product_id": product_id,
                    "price": 0, "admin": fee_val, "commission": 0,
                    "suggested_admin": suggested if suggested > 0 else 3500,
                })

    return products


def parse_pbb(ws):
    """Parse PBB sheet - property tax.
    Cols: Product ID, Tipe, Label, Suggested Admin, Service Fee Sepulsa, BPD"""
    products = []
    for row in ws.iter_rows(values_only=True):
        vals = list(row)[:7]
        first_val = vals[0]

        if isinstance(first_val, str) and ("product" in first_val.lower() or "direct" in first_val.lower()
                                           or "produk" in first_val.lower()):
            continue
        if not is_product_id(first_val):
            continue

        product_id = int(first_val)
        name = str(vals[2]).strip() if vals[2] else ""
        suggested = safe_int(vals[3])
        service_fee = safe_int(vals[4])
        bpd = str(vals[5]).strip() if vals[5] else ""

        clean_name = name.replace("\t", "").strip()

        products.append({
            "sku_code": f"ALT-PBB-{product_id}",
            "name": clean_name if clean_name else f"PBB {bpd}",
            "category": "PBB", "brand": "PBB", "type": "postpaid",
            "description": f"{clean_name} - {bpd}" if bpd else clean_name,
            "alterra_product_id": product_id,
            "price": 0, "admin": service_fee, "commission": 0,
            "suggested_admin": suggested if suggested > 0 else 3000,
        })
    return products


def parse_gas_pdam(ws):
    """Parse Gas & PDAM sheet.
    Gas cols: Product ID, Tipe, Produk, Suggested Admin, Operator, Service Fee Sepulsa
    PDAM cols: Product ID, Tipe, Area, Max Admin, Operator Code, Service Fee, Status, Source, Provinsi, Fee Partner"""
    products = []
    current_section = None

    for row in ws.iter_rows(values_only=True):
        vals = list(row)[:10]
        first_val = vals[0]

        if isinstance(first_val, str):
            lower = first_val.lower().strip()
            if lower == "gas":
                current_section = "gas"
                continue
            elif lower == "pdam":
                current_section = "pdam"
                continue
            if "product" in lower:
                continue

        if not is_product_id(first_val):
            continue

        product_id = int(first_val)

        if current_section == "gas":
            name = str(vals[2]).strip() if vals[2] else ""
            suggested = safe_int(vals[3])
            operator = str(vals[4]).strip() if vals[4] else name
            service_fee = safe_int(vals[5])
            brand = normalize_brand(operator) if operator else "PGN"
            products.append({
                "sku_code": f"ALT-GAS-{product_id}",
                "name": name, "category": "Gas PGN", "brand": brand,
                "type": "postpaid", "description": name,
                "alterra_product_id": product_id,
                "price": 0, "admin": service_fee, "commission": 0,
                "suggested_admin": suggested if suggested > 0 else 3000,
            })

        elif current_section == "pdam":
            name = str(vals[2]).strip() if vals[2] else ""
            max_admin = safe_int(vals[3])
            service_fee = safe_int(vals[5])
            status = str(vals[6]).strip().upper() if vals[6] else "OPEN"

            if status != "OPEN" and status != "":
                continue

            # commission = max_admin - service_fee (what Alterra pays us)
            commission = max(0, max_admin - service_fee) if max_admin > 0 and service_fee > 0 else 0

            products.append({
                "sku_code": f"ALT-PDAM-{product_id}",
                "name": name, "category": "PDAM", "brand": "PDAM",
                "type": "postpaid", "description": name,
                "alterra_product_id": product_id,
                "price": 0, "admin": service_fee, "commission": commission,
                "suggested_admin": max_admin if max_admin > 0 else 2500,
            })

    return products


def parse_ticket_ewallet(ws):
    """Parse Ticket & Ewallet sheet - multiple sections with different column layouts.

    Sheet structure:
      Row 1:  "Ticketing" section header
      Row 3:  Product ID header (6 cols: ID, Type, Name, Suggested Admin, Operator, Service Fee)
      Row 4:  KAI data
      Row 6:  Product ID header (3 cols: ID, Produk, Service Fee) → e-meterai
      Row 7:  E-meterai data
      Row 9:  "GRAB DRIVER" section header
      Row 10: Product ID header (7 cols with End User Price)
      Row 11+: GRAB DRIVER data
      Row 17: Product ID header (6 cols with Denom) → ewallet
      Row 18+: All other e-wallet data
    """
    products = []
    current_section = None

    for row in ws.iter_rows(values_only=True):
        vals = list(row)[:10]
        first_val = vals[0]

        # Detect section headers (non-data text rows)
        if isinstance(first_val, str):
            lower = first_val.lower().strip()

            # Named section headers
            if "ticketing" in lower:
                current_section = "ticketing"
                continue
            elif "grab driver" in lower:
                current_section = "grab_driver"
                continue

            # Column header rows - detect format by column count and content
            if "product id" in lower:
                header_str = " ".join(str(v).lower() for v in vals if v)
                populated_cols = sum(1 for v in vals[:7] if v is not None)
                if "end user price" in header_str:
                    current_section = "grab_driver"
                elif populated_cols <= 3:
                    # Compact format: ID, Produk, Fee → e-meterai
                    current_section = "emeterai"
                elif "denom" in header_str:
                    current_section = "ewallet"
                # else: keep current_section (e.g. ticketing header with 6 cols)
                continue
            continue

        if not is_product_id(first_val):
            continue

        product_id = int(first_val)

        if current_section == "ticketing":
            # Cols: ID, Type, Name, Suggested Admin, Operator, Service Fee
            name = str(vals[2]).strip() if vals[2] else ""
            suggested = safe_int(vals[3])
            operator = str(vals[4]).strip() if vals[4] else ""
            service_fee = safe_int(vals[5])
            brand = normalize_brand(operator) if operator else "LAINNYA"
            products.append({
                "sku_code": f"ALT-TIKET-{product_id}",
                "name": name, "category": "Tiket", "brand": brand,
                "type": "prepaid", "description": name,
                "alterra_product_id": product_id,
                "price": service_fee, "admin": 0, "commission": 0,
                "suggested_admin": calc_markup(service_fee) if service_fee > 0 else suggested,
            })

        elif current_section == "emeterai":
            # Cols: ID, Produk, Service Fee
            name = str(vals[1]).strip() if vals[1] else ""
            service_fee = safe_int(vals[2])
            brand = "E-METERAI" if "meterai" in name.lower() else "LAINNYA"
            products.append({
                "sku_code": f"ALT-EMONEY-{product_id}",
                "name": name, "category": "E-Money", "brand": brand,
                "type": "prepaid", "description": name,
                "alterra_product_id": product_id,
                "price": service_fee, "admin": 0, "commission": 0,
                "suggested_admin": calc_markup(service_fee),
            })

        elif current_section == "grab_driver":
            # Cols: ID, Type, Name, Denom, Operator, End User Price / Service Fee
            name = str(vals[2]).strip() if vals[2] else ""
            denom = safe_int(vals[3])
            operator = str(vals[4]).strip() if vals[4] else ""
            price = safe_int(vals[5])
            brand = normalize_brand(operator) if operator else "GRAB"
            display_name = f"Rp {format_rupiah(denom)}" if denom > 0 else name
            products.append({
                "sku_code": f"ALT-EMONEY-{brand}-{product_id}",
                "name": display_name, "category": "E-Money", "brand": brand,
                "type": "prepaid", "description": name,
                "alterra_product_id": product_id,
                "price": price, "admin": 0, "commission": 0,
                "suggested_admin": calc_markup(price),
            })

        elif current_section == "ewallet":
            # Cols: ID, Type, Name, Denom, Operator, Service Fee
            name = str(vals[2]).strip() if vals[2] else ""
            denom = safe_int(vals[3])
            operator = str(vals[4]).strip() if vals[4] else ""
            service_fee = safe_int(vals[5])
            # Some rows have operator in col 3 instead of denom
            if denom == 0 and isinstance(vals[3], str):
                operator = str(vals[3]).strip()
                service_fee = safe_int(vals[4])

            brand = normalize_brand(operator) if operator else normalize_brand(name)
            display_name = f"Rp {format_rupiah(denom)}" if denom > 0 else name
            products.append({
                "sku_code": f"ALT-EMONEY-{brand}-{product_id}",
                "name": display_name, "category": "E-Money", "brand": brand,
                "type": "prepaid", "description": name,
                "alterra_product_id": product_id,
                "price": service_fee, "admin": 0, "commission": 0,
                "suggested_admin": calc_markup(service_fee),
            })

    return products


def parse_game(ws):
    """Parse Game sheet - prepaid game top-ups.
    Cols: Product ID, Product Type, Product Name, Nominal, Publisher, Service Fee Sepulsa"""
    products = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        vals = list(row)[:7]
        product_id = vals[0]
        if not is_product_id(product_id):
            continue

        product_id = int(product_id)
        name = str(vals[2]).strip() if vals[2] else ""
        denom = safe_int(vals[3])
        publisher = str(vals[4]).strip() if vals[4] else ""
        service_fee = safe_int(vals[5])

        brand = publisher.upper().replace("TOPUP ", "").replace("TOP UP ", "").strip()
        brand = re.sub(r'[^A-Z0-9 &\-]', '', brand).strip()
        if not brand:
            brand = "LAINNYA"

        products.append({
            "sku_code": f"ALT-GAME-{product_id}",
            "name": name, "category": "Voucher Game", "brand": brand,
            "type": "prepaid", "description": name,
            "alterra_product_id": product_id,
            "price": service_fee, "admin": 0, "commission": 0,
            "suggested_admin": calc_markup(service_fee),
        })
    return products


def parse_streaming_tv(ws):
    """Parse Streaming & TV sheet - multiple sections.
    Streaming: Service Fee Sepulsa (prepaid, price=service_fee)
    Spotify/YouTube: Collection Fee for Partner as % (prepaid, price=denom, commission=denom*%)
    TV Cable Prepaid: Collection Fee for Partner as absolute (prepaid, price=denom, commission=fee)"""
    products = []
    current_section = None
    col5_is_collection_fee = False

    for row in ws.iter_rows(values_only=True):
        vals = list(row)[:10]
        first_val = vals[0]

        if isinstance(first_val, str):
            lower = first_val.lower()
            if "streaming voucher" in lower:
                current_section = "streaming"
                col5_is_collection_fee = False
                continue
            elif "tv cable prepaid" in lower:
                current_section = "tv_prepaid"
                col5_is_collection_fee = True
                continue
            elif "tv" in lower and ("postpaid" in lower or "cable" in lower or "tagihan" in lower):
                current_section = "tv_postpaid"
                col5_is_collection_fee = True
                continue
            if "product id" in lower:
                header_str = " ".join(str(v).lower() for v in vals if v)
                col5_is_collection_fee = "collection fee" in header_str
                continue
            continue

        if not is_product_id(first_val):
            continue

        product_id = int(first_val)
        tipe = str(vals[1]).strip() if vals[1] else ""
        name = str(vals[2]).strip() if vals[2] else ""
        denom = safe_int(vals[3])
        publisher = str(vals[4]).strip() if vals[4] else ""
        fee_val = safe_float(vals[5])

        brand = publisher.upper().strip()
        brand = re.sub(r'[^A-Z0-9 &\-]', '', brand).strip() or "LAINNYA"

        if col5_is_collection_fee:
            # Collection fee model: price = denom, commission calculated from fee
            if fee_val < 1:
                # Percentage: commission = denom * percentage
                commission = int(round(denom * fee_val))
            else:
                # Absolute value
                commission = int(fee_val)

            if current_section == "tv_postpaid":
                cat_prefix = "TVKABEL"
                category = "TV Kabel"
                prod_type = "postpaid"
                products.append({
                    "sku_code": f"ALT-{cat_prefix}-{product_id}",
                    "name": name, "category": category, "brand": brand,
                    "type": prod_type, "description": name,
                    "alterra_product_id": product_id,
                    "price": 0, "admin": 0, "commission": commission,
                    "suggested_admin": 3000,
                })
            else:
                # Prepaid with collection fee (Spotify, YouTube, K-Vision)
                if "paytv" in tipe.lower() or current_section == "tv_prepaid":
                    category = "TV Kabel"
                    cat_prefix = "TVKABEL"
                else:
                    category = "Streaming"
                    cat_prefix = "STREAMING"
                products.append({
                    "sku_code": f"ALT-{cat_prefix}-{product_id}",
                    "name": name, "category": category, "brand": brand,
                    "type": "prepaid", "description": name,
                    "alterra_product_id": product_id,
                    "price": denom, "admin": 0, "commission": commission,
                    "suggested_admin": calc_markup(denom),
                })
        else:
            # Service fee model: price = service_fee
            service_fee = int(fee_val)
            products.append({
                "sku_code": f"ALT-STREAMING-{product_id}",
                "name": name, "category": "Streaming", "brand": brand,
                "type": "prepaid", "description": name,
                "alterra_product_id": product_id,
                "price": service_fee, "admin": 0, "commission": 0,
                "suggested_admin": calc_markup(service_fee),
            })

    return products


def parse_voucher_deals(ws):
    """Parse Voucher Deals sheet - prepaid vouchers.
    Cols: Product ID, Product Type, Product Name, Denom, Operator, Service Fee Sepulsa"""
    products = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        vals = list(row)[:7]
        product_id = vals[0]
        if not is_product_id(product_id):
            continue

        product_id = int(product_id)
        name = str(vals[2]).strip() if vals[2] else ""
        denom = safe_int(vals[3])
        operator = str(vals[4]).strip() if vals[4] else ""
        service_fee = safe_int(vals[5])

        brand = operator.upper().replace("VOUCHER ", "").strip()
        brand = re.sub(r'[^A-Z0-9 &\-]', '', brand).strip() or "LAINNYA"
        display_name = f"Rp {format_rupiah(denom)}" if denom > 0 else name

        products.append({
            "sku_code": f"ALT-VOUCHER-{product_id}",
            "name": display_name, "category": "Voucher", "brand": brand,
            "type": "prepaid", "description": name,
            "alterra_product_id": product_id,
            "price": service_fee, "admin": 0, "commission": 0,
            "suggested_admin": calc_markup(service_fee),
        })
    return products


def parse_donation(ws):
    """Parse Donation sheet - prepaid with collection fee.
    Cols: Product ID, Tipe, Label, Denom, Operator, Collection Fee for Partner"""
    products = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        vals = list(row)[:7]
        product_id = vals[0]
        if not is_product_id(product_id):
            continue

        product_id = int(product_id)
        name = str(vals[2]).strip() if vals[2] else ""
        denom = safe_int(vals[3])
        operator = str(vals[4]).strip() if vals[4] else ""
        collection_fee = safe_int(vals[5])

        brand = operator.upper().replace("INFAQ ", "").strip()
        brand = re.sub(r'[^A-Z0-9 &\-]', '', brand).strip() or "LAINNYA"
        display_name = f"Rp {format_rupiah(denom)}" if denom > 0 else name

        products.append({
            "sku_code": f"ALT-DONASI-{product_id}",
            "name": display_name, "category": "Donasi", "brand": brand,
            "type": "prepaid", "description": name,
            "alterra_product_id": product_id,
            "price": denom, "admin": 0, "commission": collection_fee,
            "suggested_admin": calc_markup(denom),
        })
    return products


def parse_property(ws):
    """Parse Property sheet - postpaid residential payments.
    Cols: Product ID, Product Type, Product Name, Service Fee Sepulsa, Max Admin Fee"""
    products = []
    for row in ws.iter_rows(values_only=True):
        vals = list(row)[:5]
        first_val = vals[0]

        if isinstance(first_val, str):
            continue
        if not is_product_id(first_val):
            continue

        product_id = int(first_val)
        name = str(vals[2]).strip() if vals[2] else ""
        service_fee = safe_int(vals[3])
        max_admin = safe_int(vals[4])

        products.append({
            "sku_code": f"ALT-PROPERTY-{product_id}",
            "name": name, "category": "Properti", "brand": "PROPERTI",
            "type": "postpaid", "description": name,
            "alterra_product_id": product_id,
            "price": 0, "admin": service_fee, "commission": 0,
            "suggested_admin": max_admin if max_admin > 0 else 3000,
        })
    return products


def parse_edukasi(ws):
    """Parse Edukasi sheet - postpaid education payments.
    Cols: Product ID, Tipe, Label, Operator, Service Fee Sepulsa, Suggested Admin"""
    products = []
    for row in ws.iter_rows(values_only=True):
        vals = list(row)[:7]
        product_id = vals[0]
        if not is_product_id(product_id):
            continue

        product_id = int(product_id)
        name = str(vals[2]).strip() if vals[2] else ""
        operator = str(vals[3]).strip() if vals[3] else ""
        service_fee = safe_int(vals[4])
        suggested = parse_suggested_admin(vals[5])

        brand = operator.upper().strip() if operator else "LAINNYA"
        brand = re.sub(r'[^A-Z0-9 &\-]', '', brand).strip() or "LAINNYA"

        products.append({
            "sku_code": f"ALT-EDU-{product_id}",
            "name": name, "category": "Edukasi", "brand": brand,
            "type": "postpaid", "description": name,
            "alterra_product_id": product_id,
            "price": 0, "admin": service_fee, "commission": 0,
            "suggested_admin": suggested if suggested > 0 else 3000,
        })
    return products


def parse_voucher_edukasi(ws):
    """Parse Voucher Edukasi sheet - prepaid education vouchers with collection fee.
    Cols: Product ID, Product Type, Product Name, Denom, Operator, Collection Fee for Partner"""
    products = []
    prev_product_id = None
    for row in ws.iter_rows(min_row=2, values_only=True):
        vals = list(row)[:7]
        raw_id = vals[0]

        if isinstance(raw_id, str) and raw_id.startswith("="):
            if prev_product_id is not None:
                raw_id = prev_product_id + 1
            else:
                continue
        if not is_product_id(raw_id):
            continue

        product_id = int(raw_id)
        prev_product_id = product_id
        name = str(vals[2]).strip() if vals[2] else ""
        denom = safe_int(vals[3])
        operator = str(vals[4]).strip() if vals[4] else ""
        collection_fee = safe_int(vals[5])

        brand = operator.upper().strip() if operator else "LAINNYA"
        brand = re.sub(r'[^A-Z0-9 &\-]', '', brand).strip() or "LAINNYA"
        display_name = f"Rp {format_rupiah(denom)}" if denom > 0 else name

        products.append({
            "sku_code": f"ALT-VOUCHEREDU-{product_id}",
            "name": display_name, "category": "Edukasi", "brand": brand,
            "type": "prepaid", "description": name,
            "alterra_product_id": product_id,
            "price": denom, "admin": 0, "commission": collection_fee,
            "suggested_admin": calc_markup(denom),
        })
    return products


# ============================================================
# SQL Generation
# ============================================================

CATEGORIES = [
    ("Pulsa", 1), ("Listrik", 2), ("BPJS Kesehatan", 3), ("PDAM", 4),
    ("Gas PGN", 5), ("Telepon Pascabayar", 6), ("Internet", 7),
    ("TV Kabel", 8), ("Streaming", 9), ("Voucher Game", 10),
    ("E-Money", 11), ("Voucher", 12), ("Donasi", 13), ("Edukasi", 14),
    ("PBB", 15), ("Tiket", 16), ("Properti", 17),
]


def generate_sql(all_products):
    """Generate SQL migration from parsed products."""
    # Deduplicate by alterra_product_id
    seen = set()
    unique = []
    for p in all_products:
        if p["alterra_product_id"] not in seen:
            seen.add(p["alterra_product_id"])
            unique.append(p)

    # Deduplicate sku_codes
    sku_counts = {}
    for p in unique:
        sku = p["sku_code"]
        sku_counts[sku] = sku_counts.get(sku, 0) + 1

    sku_seen = {}
    for p in unique:
        sku = p["sku_code"]
        if sku_counts[sku] > 1:
            p["sku_code"] = f"{sku}-{p['alterra_product_id']}"
        if p["sku_code"] in sku_seen:
            p["sku_code"] = f"{p['sku_code']}-{p['alterra_product_id']}"
        sku_seen[p["sku_code"]] = True

    # Collect unique brands
    brands = {}
    for p in unique:
        b = p["brand"]
        if b not in brands:
            brands[b] = len(brands) + 1

    lines = []
    lines.append("-- ============================================")
    lines.append("-- Migration 000028: Reseed Alterra Pricing")
    lines.append("-- Auto-generated by scripts/parse_alterra_catalog.py")
    lines.append("-- Correct pricing: prepaid->price, postpaid->admin/commission")
    lines.append("-- Markup in products.admin, commission captured from collection_fee")
    lines.append("-- ============================================")
    lines.append("")
    lines.append("BEGIN;")
    lines.append("")

    # 1. Clean provider SKUs for Alterra only (products have FK from transactions, can't delete)
    lines.append("-- 1. Clean Alterra provider SKUs only (products preserved for FK integrity)")
    lines.append("DELETE FROM ppob_provider_skus WHERE provider_id = (SELECT id FROM ppob_providers WHERE code = 'alterra');")
    lines.append("")

    # 2. Seed master categories
    lines.append("-- 2. Seed master categories")
    cat_values = ", ".join(
        f"('{escape_sql(name)}', {order})"
        for name, order in CATEGORIES
    )
    lines.append(f"INSERT INTO product_categories (name, display_order) VALUES {cat_values}")
    lines.append("ON CONFLICT (name) DO UPDATE SET display_order = EXCLUDED.display_order;")
    lines.append("")

    # 3. Seed master brands
    lines.append("-- 3. Seed master brands")
    brand_values = ", ".join(
        f"('{escape_sql(b)}', {i})"
        for b, i in sorted(brands.items(), key=lambda x: x[1])
    )
    lines.append(f"INSERT INTO product_brands (name, display_order) VALUES {brand_values}")
    lines.append("ON CONFLICT (name) DO UPDATE SET display_order = EXCLUDED.display_order;")
    lines.append("")

    # 4. Upsert products with markup in admin field
    lines.append("-- 4. Upsert products (admin = markup for prepaid, suggested admin for postpaid)")
    for p in unique:
        product_admin = p["suggested_admin"]
        lines.append(
            f"INSERT INTO products (sku_code, name, category, brand, type, admin, commission, description, is_active) "
            f"VALUES ('{escape_sql(p['sku_code'])}', '{escape_sql(p['name'])}', '{escape_sql(p['category'])}', "
            f"'{escape_sql(p['brand'])}', '{p['type']}', {product_admin}, 0, "
            f"'{escape_sql(p['description'])}', true) "
            f"ON CONFLICT (sku_code) DO UPDATE SET name = EXCLUDED.name, category = EXCLUDED.category, "
            f"brand = EXCLUDED.brand, type = EXCLUDED.type, admin = EXCLUDED.admin, "
            f"description = EXCLUDED.description;"
        )

    lines.append("")
    lines.append("-- 5. Insert provider SKU mappings for Alterra")

    for p in unique:
        price_val = p["price"]
        admin_val = p["admin"]
        commission_val = p["commission"]

        lines.append(
            f"INSERT INTO ppob_provider_skus (provider_id, product_id, provider_sku_code, provider_product_name, price, admin, commission, is_active, is_available) "
            f"SELECT (SELECT id FROM ppob_providers WHERE code = 'alterra'), p.id, '{p['alterra_product_id']}', '{escape_sql(p['description'])}', "
            f"{price_val}, {admin_val}, {commission_val}, true, true "
            f"FROM products p WHERE p.sku_code = '{escape_sql(p['sku_code'])}' "
            f"ON CONFLICT (provider_id, provider_sku_code) DO UPDATE SET "
            f"price = EXCLUDED.price, admin = EXCLUDED.admin, commission = EXCLUDED.commission, "
            f"provider_product_name = EXCLUDED.provider_product_name, "
            f"is_active = true, is_available = true;"
        )

    lines.append("")
    lines.append("-- 6. Ensure Alterra provider is active, Digiflazz disabled")
    lines.append("UPDATE ppob_providers SET is_active = true WHERE code = 'alterra';")
    lines.append("UPDATE ppob_providers SET is_active = false WHERE code = 'digiflazz';")
    lines.append("")
    lines.append("COMMIT;")

    return "\n".join(lines)


def generate_down_sql():
    return (
        "-- Rollback: restore Digiflazz, products need manual re-seed\n"
        "UPDATE ppob_providers SET is_active = true WHERE code = 'digiflazz';\n"
        "\n"
        "-- Note: Products are not restored. Run sync worker or re-apply previous migration.\n"
    )


def main():
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True)

    all_products = []

    sheets_to_parse = [
        ("Pulsa", parse_pulsa),
        ("PLN, BPJS & Postpaid", parse_pln_bpjs_postpaid),
        ("PBB", parse_pbb),
        ("Gas & PDAM", parse_gas_pdam),
        ("Ticket & Ewallet", parse_ticket_ewallet),
        ("Game", parse_game),
        ("Streaming & TV", parse_streaming_tv),
        ("Voucher Deals", parse_voucher_deals),
        ("Donation", parse_donation),
        ("Property", parse_property),
        ("Edukasi", parse_edukasi),
        ("Voucher Edukasi", parse_voucher_edukasi),
    ]

    for sheet_name, parser_fn in sheets_to_parse:
        if sheet_name in wb.sheetnames:
            print(f"Parsing {sheet_name}...")
            results = parser_fn(wb[sheet_name])
            all_products.extend(results)
            print(f"  -> {len(results)} products")
        else:
            print(f"SKIP: Sheet '{sheet_name}' not found")

    wb.close()

    # Print summary
    categories = {}
    brands_set = set()
    types = {"prepaid": 0, "postpaid": 0}
    with_commission = 0
    for p in all_products:
        cat = p["category"]
        categories[cat] = categories.get(cat, 0) + 1
        types[p["type"]] += 1
        brands_set.add(p["brand"])
        if p["commission"] > 0:
            with_commission += 1

    print(f"\nTotal products: {len(all_products)}")
    print("By category:")
    for cat, count in sorted(categories.items()):
        print(f"  {cat}: {count}")
    print(f"By type:")
    for t, count in types.items():
        print(f"  {t}: {count}")
    print(f"Unique brands: {len(brands_set)}")
    print(f"Products with commission > 0: {with_commission}")

    # Show some pricing samples
    print("\n--- Pricing Samples ---")
    for cat in sorted(categories.keys()):
        samples = [p for p in all_products if p["category"] == cat][:2]
        for s in samples:
            print(f"  {cat} ({s['type']}): price={s['price']}, admin={s['admin']}, "
                  f"commission={s['commission']}, product_markup={s['suggested_admin']} | {s['name']}")

    # Deduplicate count
    seen = set()
    for p in all_products:
        seen.add(p["alterra_product_id"])
    print(f"\nUnique products (by alterra_product_id): {len(seen)}")

    # Generate SQL
    up_sql = generate_sql(all_products)
    down_sql = generate_down_sql()

    with open(UP_PATH, "w", encoding="utf-8") as f:
        f.write(up_sql)
    print(f"\nWritten: {UP_PATH}")

    with open(DOWN_PATH, "w", encoding="utf-8") as f:
        f.write(down_sql)
    print(f"Written: {DOWN_PATH}")


if __name__ == "__main__":
    main()
