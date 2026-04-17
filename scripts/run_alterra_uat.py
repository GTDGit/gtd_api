"""
Alterra UAT execution helper.

Runs GTD's Alterra-backed PPOB flow against the public API, then fills the
Alterra workbook using the real Alterra payloads preserved by GTD runtime
storage (PostgreSQL + Redis inquiry cache).
"""

import json
import os
import re
import shlex
import subprocess
import sys
import time
from datetime import datetime
from urllib import error as urllib_error
from urllib import request as urllib_request

import openpyxl

try:
    import requests
except ModuleNotFoundError:
    requests = None


API_BASE = "https://api.gtd.co.id"
API_KEY = "gb_live_587b6e80b00c947c02ab9da4ab96f6700af517afe6a6aedb8d8e8e09b3b43b9f"
CLIENT_ID = "ppob-id"
HEADERS = {
    "Authorization": f"Bearer {API_KEY}",
    "X-Client-Id": CLIENT_ID,
    "Content-Type": "application/json",
    "Accept": "application/json",
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/135.0.0.0 Safari/537.36"
    ),
}

XLSX_PATH = os.path.join(
    os.path.dirname(__file__), "..", "docs", "ppob", "alterra", "[GTD] Scenario UAT.xlsx"
)

SCREENSHOT_TEXT = "Aplikasi sedang dalam perubahan dari Native menjadi Webview"

SSH_TARGET = "ubuntu@15.235.143.72"
SSH_KEY = os.path.expanduser("~/.ssh/gtd")
REMOTE_BACKEND_DIR = "/home/ubuntu/backend"

PRODUCT_MAP = {
    "6": "0102014",
    "9": "0104017",
    "11": "9900011",
    "25": "0103004",
    "26": "0104012",
    "27": "9900027",
    "34": "0301007",
    "80": "0103015",
    "87": "0105020",
    "112": "9900112",
    "128": "9900128",
    "131": "9900131",
    "205": "9900205",
    "242": "9900242",
    "244": "9900244",
    "246": "9900246",
    "248": "9900248",
    "351": "9900351",
    "446": "9900446",
    "447": "9900447",
    "686": "9900686",
    "687": "9900687",
}

SHEET_FLOW = {
    "Mobile_Prepaid": "prepaid_only",
    "PLN_Prepaid": "inquiry_purchase",
    "PLN_Postpaid": "inquiry_payment",
    "BPJS_Kesehatan": "inquiry_payment",
    "BPJS_TK": "inquiry_payment",
    "PGN": "inquiry_payment",
    "PDAM": "inquiry_payment",
    "TV_Prepaid": "inquiry_payment",
    "Voucher": "prepaid_only",
    "Education": "inquiry_payment",
}

RC_SUCCESS = "00"
RC_PENDING = "10"

REF_COUNTER = int(time.time()) % 100000

DAY_NAMES_ID = {
    0: "Senin",
    1: "Selasa",
    2: "Rabu",
    3: "Kamis",
    4: "Jumat",
    5: "Sabtu",
    6: "Minggu",
}


def current_test_date():
    now = datetime.now()
    day_name = DAY_NAMES_ID[now.weekday()]
    return f"{day_name}, {now.day}-{now.month}-{now.year}"


TEST_DATE = current_test_date()
RUN_STAMP = datetime.now().strftime("%Y%m%d-%H%M%S")
OUTPUT_PATH = os.path.join(
    os.path.dirname(__file__),
    "..",
    "docs",
    "ppob",
    "alterra",
    f"[GTD] Scenario UAT - Filled {RUN_STAMP}.xlsx",
)

SCENARIO_OVERRIDES = {
    ("Mobile_Prepaid", "10"): {
        "customer_no": "0878891149161214",
    },
}

SHARED_REFERENCE_IDS = {
    "mobile_order_114": f"UAT-MOBILE-114-{RUN_STAMP}",
}


def next_ref(prefix):
    global REF_COUNTER
    REF_COUNTER += 1
    return f"UAT-{prefix}-{REF_COUNTER:04d}"


def remote_exec(script, timeout=20):
    cmd = [
        "ssh",
        "-i",
        SSH_KEY,
        SSH_TARGET,
        f"bash -lc {shlex.quote(script)}",
    ]
    result = subprocess.run(cmd, capture_output=True, text=True, timeout=timeout)
    return result.stdout.strip(), result.stderr.strip(), result.returncode


SKU_CACHE = {}


def resolve_gtd_sku(product_id):
    product_id = str(product_id).strip()
    if not product_id:
        return ""
    if product_id in SKU_CACHE:
        return SKU_CACHE[product_id]

    sql = (
        "SELECT p.sku_code "
        "FROM ppob_provider_skus pps "
        "JOIN products p ON p.id = pps.product_id "
        "JOIN ppob_providers pp ON pp.id = pps.provider_id "
        f"WHERE pp.code='alterra' AND pps.provider_sku_code='{product_id}' "
        "ORDER BY p.sku_code "
        "LIMIT 1;"
    )
    script = (
        f"cd {shlex.quote(REMOTE_BACKEND_DIR)} && "
        f"docker exec gtd-postgres psql -U gtd_user -d gtd -t -A -c {shlex.quote(sql)}"
    )
    out, _, rc = remote_exec(script, timeout=20)
    if rc == 0 and out:
        SKU_CACHE[product_id] = out.strip()
        return SKU_CACHE[product_id]

    fallback = PRODUCT_MAP.get(product_id, "")
    SKU_CACHE[product_id] = fallback
    return fallback


def query_transaction_trace(trx_id):
    if not trx_id:
        return {}

    safe_trx_id = str(trx_id).replace("'", "''")
    sql = (
        "SELECT COALESCE(json_build_object("
        "'provider_response', provider_response,"
        "'provider_initial_response', provider_initial_response,"
        "'provider_http_status', provider_http_status,"
        "'provider_initial_http_status', provider_initial_http_status,"
        "'provider_ref_id', provider_ref_id"
        ")::text, '') "
        f"FROM transactions WHERE transaction_id='{safe_trx_id}' LIMIT 1;"
    )
    script = (
        f"cd {shlex.quote(REMOTE_BACKEND_DIR)} && "
        f"docker exec gtd-postgres psql -U gtd_user -d gtd -t -A -c {shlex.quote(sql)}"
    )
    try:
        out, _, _ = remote_exec(script, timeout=60)
    except Exception:
        return {}
    if not out:
        return {}
    try:
        return json.loads(out)
    except Exception:
        return {}


def query_inquiry_cache(trx_id):
    if not trx_id:
        return {}

    script = f"""
cd {shlex.quote(REMOTE_BACKEND_DIR)}
REDIS_PASSWORD="$(grep '^REDIS_PASSWORD=' .env | head -1 | cut -d= -f2- | tr -d '\\r')"
docker exec gtd-redis redis-cli --raw --no-auth-warning -a "$REDIS_PASSWORD" GET "inquiry:trx:{trx_id}"
"""
    try:
        out, _, _ = remote_exec(script, timeout=60)
    except Exception:
        return {}
    if not out or out == "(nil)":
        return {}
    try:
        return json.loads(out)
    except Exception:
        return {}


def api_call(method, path, body=None, timeout=30):
    url = f"{API_BASE}{path}"
    if requests is not None:
        try:
            if method == "POST":
                resp = requests.post(url, headers=HEADERS, json=body, timeout=timeout)
            else:
                resp = requests.get(url, headers=HEADERS, timeout=timeout)
            try:
                return resp.status_code, resp.json(), resp.text
            except Exception:
                return resp.status_code, None, resp.text
        except requests.exceptions.Timeout:
            return 0, {"error": "timeout"}, "timeout"
        except Exception as exc:
            return 0, {"error": str(exc)}, str(exc)

    data = None
    if body is not None:
        data = json.dumps(body).encode("utf-8")
    req = urllib_request.Request(url, data=data, method=method, headers=HEADERS)
    try:
        with urllib_request.urlopen(req, timeout=timeout) as resp:
            text = resp.read().decode("utf-8")
            try:
                return resp.status, json.loads(text), text
            except Exception:
                return resp.status, None, text
    except urllib_error.HTTPError as exc:
        text = exc.read().decode("utf-8")
        try:
            return exc.code, json.loads(text), text
        except Exception:
            return exc.code, None, text
    except TimeoutError:
        return 0, {"error": "timeout"}, "timeout"
    except Exception as exc:
        return 0, {"error": str(exc)}, str(exc)


def wait_for_terminal(trx_id, max_wait=60, poll_interval=5):
    for _ in range(max_wait // poll_interval):
        time.sleep(poll_interval)
        code, resp, _ = api_call("GET", f"/v1/ppob/transaction/{trx_id}")
        if resp and resp.get("data", {}).get("status") in ("Success", "Failed"):
            return code, resp
    return api_call("GET", f"/v1/ppob/transaction/{trx_id}")[:2]


def response_status(resp):
    if not isinstance(resp, dict):
        return ""
    data = resp.get("data")
    if not isinstance(data, dict):
        return ""
    status = data.get("status")
    return "" if status in (None, "") else str(status)


def should_wait_for_terminal(trx_id, http_code, resp):
    if not trx_id or not isinstance(resp, dict):
        return False

    status = response_status(resp)
    if status in ("Success", "Failed"):
        return False

    # GTD now returns 202 for pending creates, but some older flows still
    # use 201 with a Processing status before the async callback settles.
    return http_code in (200, 201, 202) and status == "Processing"


def parse_json(text):
    if text in (None, ""):
        return None
    if isinstance(text, (dict, list)):
        return text
    try:
        return json.loads(text)
    except Exception:
        return None


def format_json_text(text):
    if text in (None, ""):
        return ""
    if isinstance(text, (dict, list)):
        return json.dumps(text, indent=2)
    obj = parse_json(text)
    if obj is None:
        return str(text)
    return json.dumps(obj, indent=2)


def extract_product_id(raw):
    if raw is None:
        return None
    s = str(raw).strip()
    match = re.match(r"\[(\d+)\]", s)
    if match:
        return match.group(1)
    if "/" in s:
        return s.split("/")[0].strip()
    s = s.replace(".0", "")
    return s if s.isdigit() else None


def extract_customer_no(raw):
    if raw is None:
        return None
    if isinstance(raw, float):
        return str(int(raw))
    return str(raw).strip()


def extract_rc(text):
    obj = parse_json(text)
    if not isinstance(obj, dict):
        return ""
    rc = obj.get("response_code")
    return "" if rc in (None, "") else str(rc)


def extract_http_status(value):
    if value in (None, "", 0):
        return ""
    return str(value)


def extract_reference_no(text):
    obj = parse_json(text)
    if not isinstance(obj, dict):
        return ""
    ref = obj.get("reference_no")
    if ref:
        return str(ref)
    data = obj.get("data")
    if isinstance(data, dict):
        for key in ("reference_no", "ref_number", "refNumber"):
            if data.get(key):
                return str(data[key])
    return ""


def extract_order_id(text):
    obj = parse_json(text)
    if not isinstance(obj, dict):
        return ""
    value = obj.get("order_id")
    return "" if value in (None, "") else str(value)


def extract_alterra_transaction_id(text):
    obj = parse_json(text)
    if not isinstance(obj, dict):
        return ""
    value = obj.get("transaction_id")
    if value in (None, "", 0, "0"):
        return ""
    return str(value)


def response_shape(text):
    obj = parse_json(text)
    if not isinstance(obj, dict):
        return "unknown"
    if "created" in obj and "changed" in obj and "status" not in obj:
        return "callback"
    if "status" in obj or "type" in obj:
        return "detail"
    return "unknown"


def status_from_rc(rc):
    if rc == RC_SUCCESS:
        return "success"
    if rc == RC_PENDING:
        return "pending"
    if rc:
        return "failed"
    return ""


def status_from_response(text):
    obj = parse_json(text)
    if not isinstance(obj, dict):
        return ""
    status = obj.get("status")
    if status:
        return str(status).strip().lower()
    return status_from_rc(extract_rc(text))


def matches_expected(expected, actual):
    if not expected or expected == "-":
        return True
    if actual == "":
        return False
    options = [part.strip() for part in str(expected).split("/")]
    return str(actual) in options


def matches_expected_http(expected, actual):
    if not expected or expected == "-":
        return True
    if actual == "":
        return False

    options = [part.strip() for part in str(expected).split("/")]
    actual_text = str(actual)
    if actual_text in options:
        return True

    # GTD now returns 202 for pending create responses that older UAT sheets
    # still document as 201.
    if actual_text == "202" and "201" in options:
        return True

    return False


def result_status_matches(expected_result, actual_status):
    if not expected_result or expected_result == "-":
        return True
    if not actual_status:
        return False
    expected_parts = [part.strip().lower() for part in str(expected_result).split("/")]
    if any(actual_status in part for part in expected_parts):
        return True
    if actual_status == "processing" and any("pending" in part for part in expected_parts):
        return True
    return False


def scenario_override(scenario):
    override = dict(SCENARIO_OVERRIDES.get((scenario["sheet"], scenario["number"]), {}))
    scenario_name = (scenario.get("name") or "").lower()

    if scenario["sheet"] in ("BPJS_Kesehatan", "BPJS_TK"):
        payment_period = "01"
        if scenario["sheet"] == "BPJS_Kesehatan" and "2 month bill" in scenario_name:
            payment_period = "02"
        elif scenario["sheet"] == "BPJS_TK" and "3 month bill" in scenario_name:
            payment_period = "03"
        elif "payment_period" in scenario_name and "00 / 13" in scenario_name:
            payment_period = "00"
        override.setdefault("data", {})
        override["data"]["payment_period"] = payment_period

    if scenario["sheet"] == "Mobile_Prepaid" and (
        "using order id 114" in scenario_name or "duplicate order id using order id 114" in scenario_name
    ):
        override["shared_reference_id"] = "mobile_order_114"

    return override


def make_alterra_request(customer_no, product_id, trx_type, order_id=None, reference_no=None, extra_data=None):
    pid = int(product_id) if product_id and str(product_id).isdigit() else product_id

    if trx_type == "inquiry":
        data = {"product_id": pid}
        if extra_data:
            data.update(extra_data)
        return json.dumps(
            {
                "customer_id": customer_no,
                "inquiry_type": "Customer_information",
                "data": data,
            },
            indent=2,
        )

    if trx_type in ("prepaid", "purchase_with_reference", "payment"):
        data = {}
        if trx_type == "prepaid":
            data = dict(extra_data or {})
        elif extra_data:
            for key in ("payment_period",):
                if key in extra_data and extra_data[key] not in (None, ""):
                    data[key] = extra_data[key]
        if trx_type in ("purchase_with_reference", "payment") and reference_no:
            data["reference_no"] = reference_no
        return json.dumps(
            {
                "customer_id": customer_no,
                "product_id": pid,
                "order_id": order_id or "",
                "data": data,
            },
            indent=2,
        )

    if trx_type == "status_check":
        return f"GET /api/v5/transaction/{order_id}" if order_id else ""

    if trx_type == "callback":
        return "Callback (push from Alterra)"

    return ""


def execute_prepaid(ref_id, sku_code, customer_no, extra_data=None):
    body = {
        "referenceId": ref_id,
        "skuCode": sku_code,
        "customerNo": customer_no,
        "type": "prepaid",
        "provider": "alterra",
    }
    if extra_data:
        body["data"] = extra_data

    code, resp, raw = api_call("POST", "/v1/ppob/transaction", body)
    result = {
        "purchase_http_code": code,
        "purchase_response": resp,
        "purchase_raw": raw,
        "request_data": extra_data or {},
        "customer_no": customer_no,
    }

    trx_id = resp.get("data", {}).get("transactionId") if resp else None
    result["transactionId"] = trx_id

    if trx_id:
        trace = query_transaction_trace(trx_id)
        result["provider_response_initial"] = trace.get("provider_initial_response") or trace.get("provider_response")
        result["provider_initial_http_status"] = trace.get("provider_initial_http_status") or trace.get("provider_http_status")
        result["provider_ref_id"] = trace.get("provider_ref_id")

    if should_wait_for_terminal(trx_id, code, resp):
        poll_code, poll_resp = wait_for_terminal(trx_id)
        result["callback_http_code"] = poll_code
        result["callback_response"] = poll_resp
        result["final_status"] = response_status(poll_resp) or "Unknown"
    else:
        result["final_status"] = response_status(resp) if resp else "Error"
        if resp:
            result["callback_http_code"] = code
            result["callback_response"] = resp

    if trx_id:
        trace = query_transaction_trace(trx_id)
        result["provider_response"] = trace.get("provider_response")
        result["provider_http_status"] = trace.get("provider_http_status")
        result["provider_ref_id"] = trace.get("provider_ref_id")

    return result


def execute_inquiry(ref_id, sku_code, customer_no, extra_data=None):
    body = {
        "referenceId": ref_id,
        "skuCode": sku_code,
        "customerNo": customer_no,
        "type": "inquiry",
        "provider": "alterra",
    }
    if extra_data:
        body["data"] = extra_data

    code, resp, raw = api_call("POST", "/v1/ppob/transaction", body)
    result = {
        "inquiry_http_code": code,
        "inquiry_response": resp,
        "inquiry_raw": raw,
        "request_data": extra_data or {},
        "customer_no": customer_no,
    }

    trx_id = resp.get("data", {}).get("transactionId") if resp else None
    result["transactionId"] = trx_id
    result["inquiry_status"] = resp.get("data", {}).get("status") if resp else "Error"

    if trx_id:
        cache = query_inquiry_cache(trx_id)
        result["provider_response"] = cache.get("providerResponse")
        result["provider_http_status"] = cache.get("providerHttpStatus")
        result["provider_ref_id"] = cache.get("providerTransactionId")
        result["cached_status"] = cache.get("status")
        result["failed_code"] = cache.get("failedCode")
        result["failed_reason"] = cache.get("failedReason")

    return result


def execute_payment(inquiry_trx_id, ref_id, sku_code, customer_no, extra_data=None):
    body = {
        "referenceId": ref_id,
        "skuCode": sku_code,
        "customerNo": customer_no,
        "type": "payment",
        "transactionId": inquiry_trx_id,
        "provider": "alterra",
    }
    if extra_data:
        body["data"] = extra_data

    code, resp, raw = api_call("POST", "/v1/ppob/transaction", body)
    result = {
        "payment_http_code": code,
        "payment_response": resp,
        "payment_raw": raw,
        "alterra_order_id": inquiry_trx_id,
        "request_data": extra_data or {},
        "customer_no": customer_no,
    }

    trx_id = resp.get("data", {}).get("transactionId") if resp else None
    result["transactionId"] = trx_id or inquiry_trx_id

    if trx_id:
        trace = query_transaction_trace(trx_id)
        result["provider_response_initial"] = trace.get("provider_initial_response") or trace.get("provider_response")
        result["provider_initial_http_status"] = trace.get("provider_initial_http_status") or trace.get("provider_http_status")
        result["provider_ref_id"] = trace.get("provider_ref_id")

    if should_wait_for_terminal(trx_id, code, resp):
        poll_code, poll_resp = wait_for_terminal(trx_id)
        result["callback_http_code"] = poll_code
        result["callback_response"] = poll_resp
        result["final_status"] = response_status(poll_resp) or "Unknown"
    else:
        result["final_status"] = response_status(resp) if resp else "Error"
        if resp:
            result["callback_http_code"] = code
            result["callback_response"] = resp

    if trx_id:
        trace = query_transaction_trace(trx_id)
        result["provider_response"] = trace.get("provider_response")
        result["provider_http_status"] = trace.get("provider_http_status")
        result["provider_ref_id"] = trace.get("provider_ref_id")

    return result


def parse_scenarios(ws, sheet_name):
    scenarios = []
    current_section = None
    current_scenario = None
    flow = SHEET_FLOW.get(sheet_name, "inquiry_payment")

    for row_idx in range(1, ws.max_row + 1):
        a = ws.cell(row_idx, 1).value
        b = ws.cell(row_idx, 2).value
        c = ws.cell(row_idx, 3).value
        d = ws.cell(row_idx, 4).value
        e = ws.cell(row_idx, 5).value
        f = ws.cell(row_idx, 6).value
        g = ws.cell(row_idx, 7).value
        h = ws.cell(row_idx, 8).value
        n = ws.cell(row_idx, 14).value

        if a and isinstance(a, str):
            a_lower = a.lower().strip()
            if "positive" in a_lower:
                current_section = "positive"
                continue
            if "negative" in a_lower and "inquiry" in a_lower:
                current_section = "negative_inquiry"
                continue
            if "negative" in a_lower and "transaction" in a_lower:
                current_section = "negative_transaction"
                continue
            if "negative" in a_lower:
                current_section = "negative"
                continue
            if "suspect" in a_lower:
                current_section = "suspect"
                continue

        if a and isinstance(a, str) and a.lower() in ("no", "note"):
            continue
        if a and isinstance(a, str) and a.startswith("Note"):
            continue

        is_scenario_start = False
        if a is not None and b is not None and c is not None:
            try:
                int(str(a).replace(".0", ""))
                is_scenario_start = True
            except Exception:
                is_scenario_start = False

        if is_scenario_start:
            current_scenario = {
                "sheet": sheet_name,
                "section": current_section,
                "flow": flow,
                "number": str(a).replace(".0", ""),
                "name": str(b),
                "product_id": extract_product_id(e),
                "customer_no": extract_customer_no(d),
                "why": str(n) if n else "",
                "steps": [],
                "rows": [],
            }
            scenarios.append(current_scenario)

        if current_scenario is None or not c:
            continue

        if is_scenario_start or a is None:
            current_scenario["steps"].append(
                {
                    "action": str(c).strip(),
                    "expected_http": str(f).replace(".0", "") if f else "",
                    "expected_rc": str(g) if g else "",
                    "expected_result": str(h) if h else "",
                    "row": row_idx,
                }
            )
            current_scenario["rows"].append(row_idx)

    return scenarios


def execute_scenario(scenario):
    override = scenario_override(scenario)
    product_id = override.get("product_id", scenario["product_id"])
    customer_no = override.get("customer_no", scenario["customer_no"])
    extra_data = override.get("data") or {}
    shared_reference_id = override.get("shared_reference_id")
    if not product_id or not customer_no:
        return {"error": f"Missing product_id={product_id} or customer_no={customer_no}"}

    sku_code = resolve_gtd_sku(product_id)
    if not sku_code:
        return {"error": f"No SKU mapping for product_id={product_id}"}

    results = {"_step_results": {}}
    current_ref_id = ""
    current_inquiry = None
    current_transaction = None

    print(
        f"  [{scenario['number']}] {scenario['name'][:60]} | "
        f"product={product_id} customer={customer_no}"
    )

    for step in scenario["steps"]:
        action = (step.get("action") or "").lower()
        row = step["row"]

        if "inquiry" in action:
            current_ref_id = SHARED_REFERENCE_IDS.get(shared_reference_id) or next_ref(
                scenario["sheet"][:6].upper().replace("_", "")
            )
            inquiry = execute_inquiry(
                current_ref_id,
                sku_code,
                customer_no,
                extra_data=extra_data,
            )
            current_inquiry = inquiry
            current_transaction = None
            results["inquiry"] = inquiry
            results["_step_results"][row] = {
                "kind": "inquiry",
                "result": inquiry,
                "product_id": product_id,
                "customer_no": customer_no,
                "request_data": extra_data,
            }
            print(
                f"    Inquiry: GTD HTTP {inquiry.get('inquiry_http_code')} "
                f"status={inquiry.get('inquiry_status')} ref={current_ref_id}"
            )
            continue

        if ("purchase" in action or "transaction" in action) and "callback" not in action and "get detail" not in action:
            if not current_ref_id:
                current_ref_id = SHARED_REFERENCE_IDS.get(shared_reference_id) or next_ref(
                    scenario["sheet"][:6].upper().replace("_", "")
                )

            if current_inquiry and current_inquiry.get("inquiry_status") == "Success" and current_inquiry.get("transactionId"):
                if scenario["flow"] == "inquiry_purchase":
                    trx = execute_prepaid(current_ref_id, sku_code, customer_no)
                    kind = "purchase"
                    print(
                        f"    Purchase: GTD HTTP {trx.get('purchase_http_code')} "
                        f"final={trx.get('final_status')} ref={current_ref_id}"
                    )
                else:
                    trx = execute_payment(
                        current_inquiry["transactionId"],
                        current_ref_id,
                        sku_code,
                        customer_no,
                        extra_data=extra_data,
                    )
                    kind = "payment"
                    print(
                        f"    Payment: GTD HTTP {trx.get('payment_http_code')} "
                        f"final={trx.get('final_status')} ref={current_ref_id}"
                    )
            elif scenario["flow"] == "prepaid_only":
                trx = execute_prepaid(
                    current_ref_id,
                    sku_code,
                    customer_no,
                    extra_data=extra_data,
                )
                kind = "purchase"
                print(
                    f"    Purchase: GTD HTTP {trx.get('purchase_http_code')} "
                    f"final={trx.get('final_status')} ref={current_ref_id}"
                )
            else:
                print("    Inquiry not successful, skipping purchase/payment")
                continue

            current_transaction = trx
            results[kind] = trx
            results["_step_results"][row] = {
                "kind": kind,
                "result": trx,
                "product_id": product_id,
                "customer_no": customer_no,
                "request_data": extra_data,
                "reference_no": extract_reference_no(
                    current_inquiry.get("provider_response") if current_inquiry else ""
                ),
            }
            continue

        if "callback" in action or "get detail" in action:
            if current_transaction is not None:
                results["_step_results"][row] = {
                    "kind": "callback",
                    "result": current_transaction,
                    "product_id": product_id,
                    "customer_no": customer_no,
                    "request_data": extra_data,
                    "reference_no": extract_reference_no(
                        current_inquiry.get("provider_response") if current_inquiry else ""
                    ),
                }

    return results


def get_step_result(results, row):
    step_results = results.get("_step_results") or {}
    return step_results.get(row)


def validate_step(step, scenario, results):
    action = step["action"].lower()
    expected_http = step["expected_http"]
    expected_rc = step["expected_rc"]
    expected_result = step["expected_result"]
    issues = []
    step_row = step.get("row")
    step_result = get_step_result(results, step_row) if step_row is not None else None

    if "inquiry" in action and step_result and step_result["kind"] == "inquiry":
        inquiry = step_result["result"]
        actual_http = extract_http_status(inquiry.get("provider_http_status"))
        actual_rc = extract_rc(inquiry.get("provider_response"))
        actual_status = status_from_response(inquiry.get("provider_response")) or (inquiry.get("inquiry_status") or "").lower()

        if not matches_expected(expected_http, actual_http):
            issues.append(f"Inquiry HTTP expected {expected_http}, got {actual_http or '-'}")
        if not matches_expected(expected_rc, actual_rc):
            issues.append(f"Inquiry RC expected {expected_rc}, got {actual_rc or '-'}")
        if not result_status_matches(expected_result, actual_status):
            issues.append(f"Inquiry result expected {expected_result}, got {actual_status or '-'}")
        if not inquiry.get("provider_response"):
            issues.append("Missing Alterra inquiry response payload")

    elif ("purchase" in action or "transaction" in action) and "callback" not in action and "get detail" not in action:
        if step_result and step_result["kind"] in ("purchase", "payment"):
            trx = step_result["result"]
            actual_http = extract_http_status(
                trx.get("purchase_http_code")
                or trx.get("payment_http_code")
                or trx.get("provider_initial_http_status")
            )
            initial_response = trx.get("provider_response_initial") or trx.get("provider_response")
            actual_rc = extract_rc(initial_response)
            actual_status = status_from_response(initial_response)
        else:
            key = "purchase" if "purchase" in results else "payment"
            if key not in results:
                return issues
            trx = results[key]
            actual_http = extract_http_status(
                trx.get("purchase_http_code")
                or trx.get("payment_http_code")
                or trx.get("provider_initial_http_status")
            )
            initial_response = trx.get("provider_response_initial") or trx.get("provider_response")
            actual_rc = extract_rc(initial_response)
            actual_status = status_from_response(initial_response)

        if not matches_expected_http(expected_http, actual_http):
            issues.append(f"Initial HTTP expected {expected_http}, got {actual_http or '-'}")
        if not matches_expected(expected_rc, actual_rc):
            issues.append(f"Initial RC expected {expected_rc}, got {actual_rc or '-'}")
        if not result_status_matches(expected_result, actual_status):
            issues.append(f"Initial result expected {expected_result}, got {actual_status or '-'}")
        if not initial_response:
            issues.append("Missing Alterra initial response payload")
        if scenario["flow"] != "prepaid_only":
            request_order_id = extract_order_id(initial_response) or trx.get("alterra_order_id") or ""
            if not request_order_id:
                issues.append("Missing actual Alterra order_id for inquiry-based purchase/payment")

    elif "callback" in action or "get detail" in action:
        if step_result and step_result["kind"] == "callback":
            trx = step_result["result"]
            final_response = trx.get("provider_response")
            actual_http = extract_http_status(trx.get("callback_http_code") or trx.get("provider_http_status")) or ("200" if final_response else "")
            actual_rc = extract_rc(final_response)
            actual_status = status_from_response(final_response) or (trx.get("final_status") or "").lower()
        else:
            key = "purchase" if "purchase" in results else "payment"
            if key not in results:
                return issues
            trx = results[key]
            final_response = trx.get("provider_response")
            actual_http = extract_http_status(trx.get("callback_http_code") or trx.get("provider_http_status")) or ("200" if final_response else "")
            actual_rc = extract_rc(final_response)
            actual_status = status_from_response(final_response) or (trx.get("final_status") or "").lower()

        if not matches_expected_http(expected_http, actual_http):
            issues.append(f"Final HTTP expected {expected_http}, got {actual_http or '-'}")
        if not matches_expected(expected_rc, actual_rc):
            issues.append(f"Final RC expected {expected_rc}, got {actual_rc or '-'}")
        if not result_status_matches(expected_result, actual_status):
            issues.append(f"Final result expected {expected_result}, got {actual_status or '-'}")
        if not final_response:
            issues.append("Missing Alterra final callback/detail payload")

    return issues


def check_result_matches(scenario, results):
    issues = []
    for step in scenario["steps"]:
        issues.extend(validate_step(step, scenario, results))
    deduped = []
    for issue in issues:
        if issue not in deduped:
            deduped.append(issue)
    return len(deduped) == 0, "; ".join(deduped)


def safe_set_cell(ws, row, col, value):
    from openpyxl.cell.cell import MergedCell

    cell = ws.cell(row, col)
    if isinstance(cell, MergedCell):
        return
    cell.value = value


def fill_scenario_rows(ws, scenario, results, matches, issues):
    for step in scenario["steps"]:
        row = step["row"]
        action = step["action"].lower()
        step_result = get_step_result(results, row)

        customer_no = scenario["customer_no"]
        request_data = {}
        reference_no = ""
        if step_result:
            customer_no = step_result.get("customer_no") or customer_no
            request_data = step_result.get("request_data") or {}
            reference_no = step_result.get("reference_no") or ""

        safe_set_cell(ws, row, 9, TEST_DATE)
        safe_set_cell(ws, row, 10, matches)
        safe_set_cell(ws, row, 13, SCREENSHOT_TEXT)

        if "inquiry" in action and step_result and step_result["kind"] == "inquiry":
            inquiry = step_result["result"]
            safe_set_cell(
                ws,
                row,
                11,
                make_alterra_request(
                    customer_no,
                    scenario["product_id"],
                    "inquiry",
                    extra_data=request_data,
                ),
            )
            safe_set_cell(ws, row, 12, format_json_text(inquiry.get("provider_response")))

        elif ("purchase" in action or "transaction" in action) and "callback" not in action and "get detail" not in action:
            if step_result and step_result["kind"] in ("purchase", "payment"):
                trx = step_result["result"]
                initial_response = trx.get("provider_response_initial") or trx.get("provider_response")
                actual_order_id = extract_order_id(initial_response) or trx.get("alterra_order_id") or trx.get("transactionId")
                if step_result["kind"] == "payment":
                    trx_type = "payment"
                elif scenario["flow"] == "inquiry_purchase":
                    trx_type = "purchase_with_reference"
                else:
                    trx_type = "prepaid"
                safe_set_cell(
                    ws,
                    row,
                    11,
                    make_alterra_request(
                        customer_no,
                        scenario["product_id"],
                        trx_type,
                        actual_order_id,
                        reference_no,
                        request_data,
                    ),
                )
                safe_set_cell(ws, row, 12, format_json_text(initial_response))

        elif "callback" in action or "get detail" in action:
            if step_result and step_result["kind"] == "callback":
                trx = step_result["result"]
                final_response = trx.get("provider_response")
                alterra_trx_id = extract_alterra_transaction_id(final_response) or trx.get("provider_ref_id") or ""
                req_type = "callback" if response_shape(final_response) == "callback" else "status_check"
                safe_set_cell(ws, row, 11, make_alterra_request(None, None, req_type, alterra_trx_id))
                safe_set_cell(ws, row, 12, format_json_text(final_response))

        if not matches and issues:
            existing = ws.cell(row, 14).value or ""
            note = f"MISMATCH: {issues}"
            safe_set_cell(ws, row, 14, f"{existing} | {note}" if existing else note)


def run_uat():
    print("=== Alterra UAT Execution ===")
    print(f"Date: {TEST_DATE}")
    print(f"API: {API_BASE}")
    print(f"Workbook: {XLSX_PATH}")
    print()

    wb = openpyxl.load_workbook(XLSX_PATH)
    total_pass = 0
    total_fail = 0
    total_error = 0
    all_issues = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n{'=' * 60}")
        print(f"Sheet: {sheet_name}")
        print(f"{'=' * 60}")

        scenarios = parse_scenarios(ws, sheet_name)
        print(f"Found {len(scenarios)} scenarios")

        for scenario in scenarios:
            try:
                results = execute_scenario(scenario)
                if "error" in results:
                    total_error += 1
                    print(f"    ERROR: {results['error']}")
                    continue

                matches, issues = check_result_matches(scenario, results)
                fill_scenario_rows(ws, scenario, results, matches, issues)

                if matches:
                    total_pass += 1
                    print("    PASS")
                else:
                    total_fail += 1
                    label = f"{sheet_name}/{scenario['number']} {scenario['name']}"
                    all_issues.append(f"{label}: {issues}")
                    print(f"    FAIL: {issues}")

                time.sleep(1)
            except Exception as exc:
                total_error += 1
                print(f"    ERROR: {exc}")

    wb.save(OUTPUT_PATH)
    print(f"\n{'=' * 60}")
    print(f"RESULTS: {total_pass} pass, {total_fail} fail, {total_error} errors")
    print(f"Output saved to: {OUTPUT_PATH}")

    if all_issues:
        print("\nIssues found:")
        for issue in all_issues:
            print(f"  - {issue}")


if __name__ == "__main__":
    sys.stdout.reconfigure(line_buffering=True)
    run_uat()
