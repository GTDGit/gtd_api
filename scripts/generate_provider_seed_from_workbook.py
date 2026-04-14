#!/usr/bin/env python3
"""
Generate workbook-based PPOB seed migration for Alterra + Kiosbank.

Source of truth:
  docs/ppob/[GTD] Provider Catalog Workbook.xlsx

This migration is intentionally destructive for testing environments:
- clears PPOB transactions/history
- clears existing products/provider mappings
- reseeds products and provider mappings from the current workbook

Usage:
  py -3 scripts/generate_provider_seed_from_workbook.py
"""

from __future__ import annotations

from collections import OrderedDict
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook


ROOT_DIR = Path(__file__).resolve().parent.parent
WORKBOOK_PATH = ROOT_DIR / "docs" / "ppob" / "[GTD] Provider Catalog Workbook.xlsx"
MIGRATIONS_DIR = ROOT_DIR / "migrations"
UP_PATH = MIGRATIONS_DIR / "000032_reset_and_seed_provider_catalog_from_workbook.up.sql"
DOWN_PATH = MIGRATIONS_DIR / "000032_reset_and_seed_provider_catalog_from_workbook.down.sql"


PREPAID_SHEET = "Prepaid"
POSTPAID_SHEET = "Postpaid"


@dataclass(frozen=True)
class ProductSeed:
    sku_code: str
    name: str
    category: str
    brand: str
    product_type: str
    description: str


@dataclass(frozen=True)
class ProviderSeed:
    provider_code: str
    sku_code: str
    provider_sku_code: str
    provider_product_name: str
    price: int
    admin: int
    commission: int


def sql_string(value: str) -> str:
    return "'" + value.replace("'", "''") + "'"


def sql_int(value: int) -> str:
    return str(int(value))


def trim(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def to_int(value) -> int:
    if value in (None, ""):
        return 0
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, (int, float)):
        return int(value)
    text = trim(value).replace(",", "").replace(".", "")
    if not text:
        return 0
    return int(text)


def chunked(items: list[str], size: int = 500) -> Iterable[list[str]]:
    for i in range(0, len(items), size):
        yield items[i:i + size]


def parse_workbook() -> tuple[list[ProductSeed], list[ProviderSeed], list[tuple[str, int]], list[tuple[str, int]]]:
    wb = load_workbook(WORKBOOK_PATH, data_only=True)

    categories: "OrderedDict[str, int]" = OrderedDict()
    brands: "OrderedDict[str, int]" = OrderedDict()
    products: "OrderedDict[str, ProductSeed]" = OrderedDict()
    provider_rows: list[ProviderSeed] = []

    def remember_category(name: str) -> None:
        if name and name not in categories:
            categories[name] = len(categories) + 1

    def remember_brand(name: str) -> None:
        if name and name not in brands:
            brands[name] = len(brands) + 1

    def upsert_product(seed: ProductSeed) -> None:
        existing = products.get(seed.sku_code)
        if existing and existing != seed:
            raise ValueError(
                f"Conflicting product seed for {seed.sku_code}: {existing!r} vs {seed!r}"
            )
        products[seed.sku_code] = seed

    def add_provider(seed: ProviderSeed) -> None:
        if not seed.provider_sku_code:
            raise ValueError(f"Missing provider_sku_code for {seed.sku_code} ({seed.provider_code})")
        provider_rows.append(seed)

    prepaid = wb[PREPAID_SHEET]
    prepaid_headers = {trim(prepaid.cell(2, c).value): c for c in range(1, prepaid.max_column + 1)}

    for row_idx in range(3, prepaid.max_row + 1):
        sku_code = trim(prepaid.cell(row_idx, prepaid_headers["Kode Produk GTD"]).value)
        if not sku_code:
            continue

        category = trim(prepaid.cell(row_idx, prepaid_headers["Jenis Produk"]).value)
        brand = trim(prepaid.cell(row_idx, prepaid_headers["Brand"]).value)
        name = trim(prepaid.cell(row_idx, prepaid_headers["Nama SKU"]).value)
        status = trim(prepaid.cell(row_idx, prepaid_headers["Status"]).value).lower()

        if status not in {"both", "kiosbank_only", "alterra_only"}:
            raise ValueError(f"Unexpected status {status!r} in {PREPAID_SHEET} row {row_idx}")

        remember_category(category)
        remember_brand(brand)
        upsert_product(ProductSeed(
            sku_code=sku_code,
            name=name,
            category=category,
            brand=brand,
            product_type="prepaid",
            description=name,
        ))

        if status in {"both", "kiosbank_only"}:
            add_provider(ProviderSeed(
                provider_code="kiosbank",
                sku_code=sku_code,
                provider_sku_code=trim(prepaid.cell(row_idx, prepaid_headers["KB: Kode Produk"]).value),
                provider_product_name=trim(prepaid.cell(row_idx, prepaid_headers["KB: Nama Produk"]).value) or name,
                price=to_int(prepaid.cell(row_idx, prepaid_headers["KB: Harga Jual (Rp)"]).value),
                admin=0,
                commission=to_int(prepaid.cell(row_idx, prepaid_headers["KB: Fee Mitra (Rp)"]).value),
            ))

        if status in {"both", "alterra_only"}:
            add_provider(ProviderSeed(
                provider_code="alterra",
                sku_code=sku_code,
                provider_sku_code=trim(prepaid.cell(row_idx, prepaid_headers["ALT: Product ID"]).value),
                provider_product_name=trim(prepaid.cell(row_idx, prepaid_headers["ALT: Nama Produk"]).value) or name,
                price=to_int(prepaid.cell(row_idx, prepaid_headers["ALT: Harga / Service Fee (Rp)"]).value),
                admin=0,
                commission=to_int(prepaid.cell(row_idx, prepaid_headers["ALT: Commission (Rp)"]).value),
            ))

    postpaid = wb[POSTPAID_SHEET]
    postpaid_headers = {trim(postpaid.cell(2, c).value): c for c in range(1, postpaid.max_column + 1)}

    for row_idx in range(3, postpaid.max_row + 1):
        sku_code = trim(postpaid.cell(row_idx, postpaid_headers["Kode Produk GTD"]).value)
        if not sku_code:
            continue

        category = trim(postpaid.cell(row_idx, postpaid_headers["Jenis Produk"]).value)
        brand = trim(postpaid.cell(row_idx, postpaid_headers["Brand"]).value)
        name = trim(postpaid.cell(row_idx, postpaid_headers["Nama SKU / Biller"]).value)
        status = trim(postpaid.cell(row_idx, postpaid_headers["Status"]).value).lower()

        if status not in {"both", "kiosbank_only", "alterra_only"}:
            raise ValueError(f"Unexpected status {status!r} in {POSTPAID_SHEET} row {row_idx}")

        remember_category(category)
        remember_brand(brand)
        upsert_product(ProductSeed(
            sku_code=sku_code,
            name=name,
            category=category,
            brand=brand,
            product_type="postpaid",
            description=name,
        ))

        if status in {"both", "kiosbank_only"}:
            add_provider(ProviderSeed(
                provider_code="kiosbank",
                sku_code=sku_code,
                provider_sku_code=trim(postpaid.cell(row_idx, postpaid_headers["KB: Kode Produk"]).value),
                provider_product_name=trim(postpaid.cell(row_idx, postpaid_headers["KB: Nama Biller"]).value) or name,
                price=0,
                admin=to_int(postpaid.cell(row_idx, postpaid_headers["KB: Biaya Admin (Rp)"]).value),
                commission=to_int(postpaid.cell(row_idx, postpaid_headers["KB: Fee Mitra (Rp)"]).value),
            ))

        if status in {"both", "alterra_only"}:
            add_provider(ProviderSeed(
                provider_code="alterra",
                sku_code=sku_code,
                provider_sku_code=trim(postpaid.cell(row_idx, postpaid_headers["ALT: Product ID"]).value),
                provider_product_name=trim(postpaid.cell(row_idx, postpaid_headers["ALT: Nama Produk"]).value) or name,
                price=0,
                admin=to_int(postpaid.cell(row_idx, postpaid_headers["ALT: Service Fee (Rp)"]).value),
                commission=to_int(postpaid.cell(row_idx, postpaid_headers["ALT: Collection Fee Partner (Rp)"]).value),
            ))

    return (
        list(products.values()),
        provider_rows,
        list(categories.items()),
        list(brands.items()),
    )


def build_values_rows(rows: list[str], indent: str = "    ") -> list[str]:
    lines: list[str] = []
    for chunk in chunked(rows):
        lines.append("INSERT INTO")
        lines.extend(chunk)
    return lines


def insert_statement(header: str, rows: list[str]) -> list[str]:
    lines: list[str] = []
    for chunk in chunked(rows):
        lines.append(header)
        for index, row in enumerate(chunk):
            suffix = "," if index < len(chunk) - 1 else ";"
            lines.append(f"    {row}{suffix}")
        lines.append("")
    return lines


def generate_up_sql(
    products: list[ProductSeed],
    provider_rows: list[ProviderSeed],
    categories: list[tuple[str, int]],
    brands: list[tuple[str, int]],
) -> str:
    product_rows = [
        "("
        + ", ".join(
            [
                sql_string(p.sku_code),
                sql_string(p.name),
                sql_string(p.category),
                sql_string(p.brand),
                sql_string(p.product_type),
                sql_string(p.description),
            ]
        )
        + ")"
        for p in products
    ]

    provider_value_rows = [
        "("
        + ", ".join(
            [
                sql_string(p.provider_code),
                sql_string(p.sku_code),
                sql_string(p.provider_sku_code),
                sql_string(p.provider_product_name),
                sql_int(p.price),
                sql_int(p.admin),
                sql_int(p.commission),
            ]
        )
        + ")"
        for p in provider_rows
    ]

    category_rows = [
        "(" + ", ".join([sql_string(name), sql_int(display_order)]) + ")"
        for name, display_order in categories
    ]
    brand_rows = [
        "(" + ", ".join([sql_string(name), sql_int(display_order)]) + ")"
        for name, display_order in brands
    ]

    lines = [
        "-- ============================================",
        "-- Migration 000032: Reset and seed provider catalog from workbook",
        "-- Source: docs/ppob/[GTD] Provider Catalog Workbook.xlsx",
        "-- WARNING: destructive reset for PPOB testing environments",
        "-- Postpaid mapping:",
        "--   Kiosbank  : admin = biaya admin, commission = fee mitra",
        "--   Alterra   : admin = service fee, commission = collection fee",
        "-- ============================================",
        "",
        "BEGIN;",
        "",
        "-- Allow duplicated provider_sku_code for Kiosbank shared SKU cases such as XL/AXIS.",
        "ALTER TABLE ppob_provider_skus",
        "    DROP CONSTRAINT IF EXISTS ppob_provider_skus_provider_id_provider_sku_code_key;",
        "",
        "-- Clear PPOB history for testing reset.",
        "DELETE FROM ppob_provider_callbacks;",
        "DELETE FROM digiflazz_callbacks;",
        "DELETE FROM transaction_logs;",
        "DELETE FROM transactions;",
        "DELETE FROM ppob_provider_health;",
        "DELETE FROM ppob_provider_skus;",
        "DELETE FROM skus;",
        "DELETE FROM products;",
        "DELETE FROM product_categories;",
        "DELETE FROM product_brands;",
        "",
        "CREATE TEMP TABLE tmp_seed_products (",
        "    sku_code VARCHAR(50) PRIMARY KEY,",
        "    name VARCHAR(100) NOT NULL,",
        "    category VARCHAR(100) NOT NULL,",
        "    brand VARCHAR(100) NOT NULL,",
        "    product_type product_type NOT NULL,",
        "    description TEXT NOT NULL",
        ") ON COMMIT DROP;",
        "",
    ]

    lines.extend(insert_statement(
        "INSERT INTO tmp_seed_products (sku_code, name, category, brand, product_type, description) VALUES",
        product_rows,
    ))

    lines.extend([
        "CREATE TEMP TABLE tmp_seed_provider_skus (",
        "    provider_code VARCHAR(20) NOT NULL,",
        "    sku_code VARCHAR(50) NOT NULL,",
        "    provider_sku_code VARCHAR(100) NOT NULL,",
        "    provider_product_name VARCHAR(200) NOT NULL,",
        "    price INT NOT NULL,",
        "    admin INT NOT NULL,",
        "    commission INT NOT NULL",
        ") ON COMMIT DROP;",
        "",
    ])
    lines.extend(insert_statement(
        "INSERT INTO tmp_seed_provider_skus (provider_code, sku_code, provider_sku_code, provider_product_name, price, admin, commission) VALUES",
        provider_value_rows,
    ))

    lines.extend([
        "CREATE TEMP TABLE tmp_seed_categories (",
        "    name VARCHAR(100) PRIMARY KEY,",
        "    display_order INT NOT NULL",
        ") ON COMMIT DROP;",
        "",
    ])
    lines.extend(insert_statement(
        "INSERT INTO tmp_seed_categories (name, display_order) VALUES",
        category_rows,
    ))

    lines.extend([
        "CREATE TEMP TABLE tmp_seed_brands (",
        "    name VARCHAR(100) PRIMARY KEY,",
        "    display_order INT NOT NULL",
        ") ON COMMIT DROP;",
        "",
    ])
    lines.extend(insert_statement(
        "INSERT INTO tmp_seed_brands (name, display_order) VALUES",
        brand_rows,
    ))

    lines.extend([
        "INSERT INTO product_categories (name, display_order)",
        "SELECT name, display_order FROM tmp_seed_categories",
        "ORDER BY display_order, name;",
        "",
        "INSERT INTO product_brands (name, display_order)",
        "SELECT name, display_order FROM tmp_seed_brands",
        "ORDER BY display_order, name;",
        "",
        "INSERT INTO products (sku_code, name, category, brand, type, admin, commission, description, is_active)",
        "SELECT",
        "    sku_code,",
        "    name,",
        "    category,",
        "    brand,",
        "    product_type,",
        "    0 AS admin,",
        "    0 AS commission,",
        "    description,",
        "    true AS is_active",
        "FROM tmp_seed_products",
        "ORDER BY sku_code;",
        "",
        "INSERT INTO ppob_provider_skus (",
        "    provider_id,",
        "    product_id,",
        "    provider_sku_code,",
        "    provider_product_name,",
        "    price,",
        "    admin,",
        "    commission,",
        "    is_active,",
        "    is_available",
        ")",
        "SELECT",
        "    pr.id,",
        "    p.id,",
        "    s.provider_sku_code,",
        "    s.provider_product_name,",
        "    s.price,",
        "    s.admin,",
        "    s.commission,",
        "    true,",
        "    true",
        "FROM tmp_seed_provider_skus s",
        "JOIN ppob_providers pr ON pr.code = s.provider_code",
        "JOIN products p ON p.sku_code = s.sku_code",
        "ORDER BY p.sku_code, pr.code, s.provider_sku_code;",
        "",
        "COMMIT;",
        "",
    ])

    return "\n".join(lines)


def generate_down_sql() -> str:
    return "\n".join([
        "-- ============================================",
        "-- Migration 000032 DOWN: Remove workbook-based PPOB seed",
        "-- WARNING: destructive reset for PPOB testing environments",
        "-- ============================================",
        "",
        "BEGIN;",
        "",
        "DELETE FROM ppob_provider_callbacks;",
        "DELETE FROM digiflazz_callbacks;",
        "DELETE FROM transaction_logs;",
        "DELETE FROM transactions;",
        "DELETE FROM ppob_provider_health;",
        "DELETE FROM ppob_provider_skus;",
        "DELETE FROM skus;",
        "DELETE FROM products;",
        "DELETE FROM product_categories;",
        "DELETE FROM product_brands;",
        "",
        "DO $$",
        "BEGIN",
        "    IF NOT EXISTS (",
        "        SELECT 1",
        "        FROM pg_constraint",
        "        WHERE conname = 'ppob_provider_skus_provider_id_provider_sku_code_key'",
        "    ) THEN",
        "        ALTER TABLE ppob_provider_skus",
        "            ADD CONSTRAINT ppob_provider_skus_provider_id_provider_sku_code_key",
        "            UNIQUE (provider_id, provider_sku_code);",
        "    END IF;",
        "END $$;",
        "",
        "COMMIT;",
        "",
    ])


def main() -> None:
    products, provider_rows, categories, brands = parse_workbook()

    if not products:
        raise ValueError("No products found in workbook")
    if not provider_rows:
        raise ValueError("No provider mappings found in workbook")

    UP_PATH.write_text(
        generate_up_sql(products, provider_rows, categories, brands),
        encoding="utf-8",
    )
    DOWN_PATH.write_text(generate_down_sql(), encoding="utf-8")

    print(f"Wrote {UP_PATH}")
    print(f"Wrote {DOWN_PATH}")
    print(f"Products: {len(products)}")
    print(f"Provider mappings: {len(provider_rows)}")
    print(f"Categories: {len(categories)}")
    print(f"Brands: {len(brands)}")


if __name__ == "__main__":
    main()
