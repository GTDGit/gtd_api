import importlib.util
import json
import pathlib
import unittest

import openpyxl


MODULE_PATH = pathlib.Path(__file__).with_name("run_alterra_uat.py")
SPEC = importlib.util.spec_from_file_location("run_alterra_uat", MODULE_PATH)
MODULE = importlib.util.module_from_spec(SPEC)
SPEC.loader.exec_module(MODULE)


class AlterraUATScriptTests(unittest.TestCase):
    def test_should_wait_for_terminal_for_processing_202(self):
        should_wait = MODULE.should_wait_for_terminal(
            "GRB-20260417-000001",
            202,
            {"data": {"status": "Processing"}},
        )

        self.assertTrue(should_wait)

    def test_should_not_wait_for_terminal_for_final_status(self):
        should_wait = MODULE.should_wait_for_terminal(
            "GRB-20260417-000001",
            503,
            {"data": {"status": "Failed"}},
        )

        self.assertFalse(should_wait)

    def test_matches_expected_http_accepts_pending_202_for_legacy_201(self):
        self.assertTrue(MODULE.matches_expected_http("201", "202"))

    def test_scenario_override_reuses_mobile_duplicate_reference(self):
        override = MODULE.scenario_override(
            {
                "sheet": "Mobile_Prepaid",
                "number": "7",
                "name": "Failed Transaction, Duplicate Order ID Using Order ID 114",
            }
        )

        self.assertEqual(override["shared_reference_id"], "mobile_order_114")

    def test_bpjs_inquiry_request_includes_payment_period(self):
        payload = json.loads(
            MODULE.make_alterra_request(
                "0000001430071801",
                "34",
                "inquiry",
                extra_data={"payment_period": "02"},
            )
        )

        self.assertEqual(payload["data"]["product_id"], 34)
        self.assertEqual(payload["data"]["payment_period"], "02")

    def test_make_alterra_request_payment_includes_reference_no_and_payment_period(self):
        payload = json.loads(
            MODULE.make_alterra_request(
                "01428800700",
                "25",
                "payment",
                order_id="TRX-123",
                reference_no="REF-999",
                extra_data={"payment_period": "02"},
            )
        )

        self.assertEqual(payload["order_id"], "TRX-123")
        self.assertEqual(payload["data"]["reference_no"], "REF-999")
        self.assertEqual(payload["data"]["payment_period"], "02")

    def test_scenario_override_applies_vendor_number_for_mobile_restriction(self):
        override = MODULE.scenario_override(
            {
                "sheet": "Mobile_Prepaid",
                "number": "10",
            }
        )

        self.assertEqual(override["customer_no"], "0878891149161214")

    def test_scenario_override_applies_bpjs_payment_period(self):
        override = MODULE.scenario_override(
            {
                "sheet": "BPJS_Kesehatan",
                "number": "2",
                "name": 'Success 2 month bill ("payment_period" : "02")',
            }
        )

        self.assertEqual(override["data"]["payment_period"], "02")

    def test_scenario_override_applies_bpjs_tk_payment_period(self):
        override = MODULE.scenario_override(
            {
                "sheet": "BPJS_TK",
                "number": "2",
                "name": 'Success 3 month bill ("payment_period" : "03")',
            }
        )

        self.assertEqual(override["data"]["payment_period"], "03")

    def test_scenario_override_does_not_apply_bpjs_period_by_reused_number_alone(self):
        override = MODULE.scenario_override(
            {
                "sheet": "BPJS_Kesehatan",
                "number": "2",
                "name": "Product Issue",
            }
        )

        self.assertEqual(override["data"]["payment_period"], "01")

    def test_validate_step_uses_alterra_http_and_rc(self):
        step = {
            "action": "Purchase",
            "expected_http": "406",
            "expected_rc": "20",
            "expected_result": "Failed",
        }
        scenario = {"flow": "prepaid_only"}
        results = {
            "purchase": {
                "provider_initial_http_status": 406,
                "provider_response_initial": json.dumps(
                    {
                        "order_id": "ORD-406",
                        "response_code": "20",
                        "status": "Failed",
                    }
                ),
            }
        }

        self.assertEqual(MODULE.validate_step(step, scenario, results), [])

    def test_validate_step_requires_actual_order_id_for_inquiry_based_purchase(self):
        step = {
            "action": "Purchase",
            "expected_http": "201",
            "expected_rc": "10",
            "expected_result": "Pending",
        }
        scenario = {"flow": "inquiry_purchase"}
        results = {
            "purchase": {
                "provider_initial_http_status": 201,
                "provider_response_initial": json.dumps(
                    {
                        "order_id": "",
                        "response_code": "10",
                        "status": "Pending",
                    }
                ),
            }
        }

        issues = MODULE.validate_step(step, scenario, results)

        self.assertIn(
            "Missing actual Alterra order_id for inquiry-based purchase/payment",
            issues,
        )

    def test_fill_scenario_rows_never_logs_transaction_zero(self):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        scenario = {
            "customer_no": "01428800700",
            "product_id": "25",
            "flow": "inquiry_purchase",
            "steps": [{"action": "Get detail/Callback", "row": 1}],
        }
        results = {
            "purchase": {
                "provider_response": json.dumps(
                    {
                        "transaction_id": 0,
                        "response_code": "10",
                        "status": "Pending",
                    }
                )
            }
        }

        MODULE.fill_scenario_rows(sheet, scenario, results, True, "")

        self.assertNotIn("/transaction/0", sheet.cell(1, 11).value or "")

    def test_format_json_text_accepts_dict_payloads(self):
        formatted = MODULE.format_json_text({"response_code": "10", "status": "Pending"})

        self.assertIn('"response_code": "10"', formatted)
        self.assertIn('"status": "Pending"', formatted)

    def test_fill_scenario_rows_uses_step_specific_results_for_multi_cycle_scenario(self):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        scenario = {
            "customer_no": "211001025251",
            "product_id": "80",
            "flow": "inquiry_payment",
            "steps": [
                {"action": "Inquiry (6 month)", "row": 1},
                {"action": "Purchase (4 month)", "row": 2},
                {"action": "Get detail/Callback", "row": 3},
                {"action": "Inquiry (2 month)", "row": 4},
                {"action": "Purchase (2 month)", "row": 5},
                {"action": "Get detail/Callback", "row": 6},
            ],
        }
        results = {
            "_step_results": {
                1: {
                    "kind": "inquiry",
                    "result": {
                        "provider_response": json.dumps(
                            {"reference_no": "REF-6M", "response_code": "00", "status": "Success"}
                        )
                    },
                    "customer_no": "211001025251",
                    "request_data": {},
                },
                2: {
                    "kind": "payment",
                    "result": {
                        "alterra_order_id": "INQ-6M",
                        "provider_response_initial": json.dumps(
                            {"order_id": "INQ-6M", "response_code": "10", "status": "Pending"}
                        ),
                        "provider_response": json.dumps(
                            {"transaction_id": 600001, "response_code": "00", "status": "Success"}
                        ),
                    },
                    "customer_no": "211001025251",
                    "request_data": {},
                    "reference_no": "REF-6M",
                },
                3: {
                    "kind": "callback",
                    "result": {
                        "provider_response": json.dumps(
                            {"transaction_id": 600001, "response_code": "00", "status": "Success"}
                        )
                    },
                    "customer_no": "211001025251",
                    "request_data": {},
                    "reference_no": "REF-6M",
                },
                4: {
                    "kind": "inquiry",
                    "result": {
                        "provider_response": json.dumps(
                            {"reference_no": "REF-2M", "response_code": "00", "status": "Success"}
                        )
                    },
                    "customer_no": "211001025251",
                    "request_data": {},
                },
                5: {
                    "kind": "payment",
                    "result": {
                        "alterra_order_id": "INQ-2M",
                        "provider_response_initial": json.dumps(
                            {"order_id": "INQ-2M", "response_code": "10", "status": "Pending"}
                        ),
                        "provider_response": json.dumps(
                            {"transaction_id": 600002, "response_code": "00", "status": "Success"}
                        ),
                    },
                    "customer_no": "211001025251",
                    "request_data": {},
                    "reference_no": "REF-2M",
                },
                6: {
                    "kind": "callback",
                    "result": {
                        "provider_response": json.dumps(
                            {"transaction_id": 600002, "response_code": "00", "status": "Success"}
                        )
                    },
                    "customer_no": "211001025251",
                    "request_data": {},
                    "reference_no": "REF-2M",
                },
            }
        }

        MODULE.fill_scenario_rows(sheet, scenario, results, True, "")

        self.assertIn('"reference_no": "REF-6M"', sheet.cell(1, 12).value)
        self.assertIn('"order_id": "INQ-6M"', sheet.cell(2, 11).value)
        self.assertIn("/transaction/600001", sheet.cell(3, 11).value)
        self.assertIn('"reference_no": "REF-2M"', sheet.cell(4, 12).value)
        self.assertIn('"order_id": "INQ-2M"', sheet.cell(5, 11).value)
        self.assertIn("/transaction/600002", sheet.cell(6, 11).value)


if __name__ == "__main__":
    unittest.main()
